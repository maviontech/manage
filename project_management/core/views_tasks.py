
import csv, io, datetime, os
from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponse
from django.views.decorators.http import require_POST
from django.views.decorators.http import require_GET, require_POST
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.shortcuts import get_object_or_404
from django.conf import settings
from django.core.files.storage import default_storage

# helper: get connection for current tenant
from .db_helpers import get_tenant_conn, get_visible_task_user_ids, get_tenant_work_types, resolve_tenant_key_from_request
from .notifications import NotificationManager
import json

# PDF generation
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.pdfgen import canvas


# ==============================
#  HELPER FUNCTION FOR ATTACHMENTS
# ==============================
def save_task_attachments(request, task_id, cur, created_by):
    """
    Helper function to save task attachments to disk and database
    
    Args:
        request: Django request object containing FILES
        task_id: The ID of the task to attach files to
        cur: Database cursor for executing queries
        created_by: User ID of the person uploading files
    
    Returns:
        int: Number of files successfully saved
    """
    attachments = request.FILES.getlist('attachments')
    if not attachments:
        return 0
    
    # Create attachments directory if it doesn't exist
    attachments_dir = os.path.join(settings.MEDIA_ROOT, 'task_attachments')
    os.makedirs(attachments_dir, exist_ok=True)
    
    saved_count = 0
    for uploaded_file in attachments:
        try:
            # Generate unique filename with timestamp
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_filename = f"{task_id}_{timestamp}_{uploaded_file.name}"
            file_path = os.path.join(attachments_dir, safe_filename)
            
            # Save the file to disk
            with open(file_path, 'wb+') as destination:
                for chunk in uploaded_file.chunks():
                    destination.write(chunk)
            
            # Store relative path for database
            relative_path = os.path.join('task_attachments', safe_filename)
            
            # Save attachment info to database
            cur.execute("""
                INSERT INTO task_attachments 
                (task_id, file_name, file_path, file_size, file_type, uploaded_by)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (
                task_id,
                uploaded_file.name,
                relative_path,
                uploaded_file.size,
                uploaded_file.content_type,
                created_by
            ))
            # Log attachment upload to activity timeline
            try:
                cur.execute(
                    "INSERT INTO activity_log (entity_type, entity_id, action, performed_by, timestamp) VALUES (%s,%s,%s,%s,NOW())",
                    ("task", task_id, f"Added attachment {uploaded_file.name}", created_by),
                )
            except Exception:
                pass
            saved_count += 1
        except Exception as e:
            # Log error but continue processing other files
            print(f"Error saving attachment {uploaded_file.name}: {e}")
            continue
    
    return saved_count


# ==============================
#  CREATE TASK  (GET / POST)
# ==============================
def create_task_view(request):
    """
    Creates a new task and saves to DB. Supports member/team polymorphic assignment.
    Requires `assigned_type` column in `tasks` table (ENUM('member','team')).
    """
    conn = get_tenant_conn(request)
    cur = conn.cursor()

    if request.method == "POST":
        data = request.POST
        project_id = data.get("project_id") or None
        subproject_id = data.get("subproject_id") or None
        title = data.get("title")
        description = data.get("description")
        due_date = data.get("due_date") or None
        closure_date = data.get("closure_date") or None
        priority = data.get("priority") or "Normal"
        status = data.get("status") or "Open"
        work_type = data.get("work_type") or "Task"  # NEW: Get work type from form
        created_by = request.session.get("user_id")
        
        # Capture system information
        si_browser = data.get("si_browser") or None
        si_resolution = data.get("si_resolution") or None
        si_os = data.get("si_os") or None
        si_timestamp = data.get("si_timestamp") or None

        # --- NEW FIX ---
        assigned_raw = data.get("assigned_to") or None
        assigned_to, assigned_type = None, None
        if assigned_raw:
            if ":" in assigned_raw:
                assigned_type, assigned_to = assigned_raw.split(":", 1)
            else:
                assigned_type, assigned_to = "member", assigned_raw

        # --- INSERT ---
        cur.execute(
            """INSERT INTO tasks
               (project_id, subproject_id, title, description, status, priority,
                assigned_to, assigned_type, created_by, due_date, closure_date, work_type, 
                si_browser, si_resolution, si_os, si_timestamp, created_at)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())""",
            (
                project_id,
                subproject_id,
                title,
                description,
                status,
                priority,
                assigned_to,
                assigned_type,
                created_by,
                due_date,
                closure_date,
                work_type,
                si_browser,
                si_resolution,
                si_os,
                si_timestamp,
            ),
        )
        task_id = cur.lastrowid

        # Log activity
        cur.execute(
            "INSERT INTO activity_log (entity_type, entity_id, action, performed_by) VALUES (%s,%s,%s,%s)",
            ("task", task_id, "created", created_by),
        )

        # Create notification if task is assigned to a member
        if assigned_type == "member" and assigned_to:
            # Get creator name
            cur.execute("SELECT CONCAT(first_name, ' ', last_name) as name FROM members WHERE id=%s", (created_by,))
            creator = cur.fetchone()
            creator_name = creator['name'] if creator else 'Someone'
            
            cur.execute("""
                INSERT INTO notifications (user_id, title, message, type, link)
                VALUES (%s, %s, %s, %s, %s)
            """, (
                assigned_to,
                "New Task Assigned",
                f"{creator_name} assigned you to '{title}'",
                "task",
                f"/tasks/{task_id}/view/"
            ))
        
        # Create notification for team members if assigned to a team
        elif assigned_type == "team" and assigned_to:
            cur.execute("SELECT CONCAT(first_name, ' ', last_name) as name FROM members WHERE id=%s", (created_by,))
            creator = cur.fetchone()
            creator_name = creator['name'] if creator else 'Someone'
            
            # Get all team members
            cur.execute("SELECT member_id FROM team_memberships WHERE team_id=%s", (assigned_to,))
            team_members = cur.fetchall()
            
            for member in team_members:
                cur.execute("""
                    INSERT INTO notifications (user_id, title, message, type, link)
                    VALUES (%s, %s, %s, %s, %s)
                """, (
                    member['member_id'],
                    "New Team Task Assigned",
                    f"{creator_name} assigned a task to your team: '{title}'",
                    "task",
                    f"/tasks/{task_id}/view/"
                ))

        # Handle file attachments using helper function
        save_task_attachments(request, task_id, cur, created_by)

        conn.commit()
        cur.close()
        return redirect("task_board")

    # --- GET ---
    cur.execute("SELECT id, name FROM projects ORDER BY name")
    projects = cur.fetchall()
    cur.execute("SELECT id, email, first_name, last_name FROM members ORDER BY first_name")
    members = cur.fetchall()
    cur.execute("SELECT id, name FROM teams ORDER BY name")
    teams = cur.fetchall()
    
    # Get tenant-specific work types from configuration
    work_types = get_tenant_work_types(request)
    
    cur.close()

    return render(
        request,
        "core/create_task.html",
        {
            "projects": projects, 
            "members": members, 
            "teams": teams, 
            "page": "task_create",
            "work_types": work_types
        },
    )


# ==============================
#  MY TASKS
# ==============================
def my_tasks_view(request):

    conn = get_tenant_conn(request)
    cur = conn.cursor()

    # Ensure user_id is set in session (should be set at login)
    user_id = request.session.get("user_id")
    if not user_id:
        # Try to resolve from email if available
        email = request.session.get("auth_email")
        if email:
            cur.execute("SELECT id FROM members WHERE email = %s", (email,))
            row = cur.fetchone()
            if row:
                user_id = row["id"]
                request.session["user_id"] = user_id
    if not user_id:
        return redirect("login")

    # Get visible task user IDs based on Alex Carter visibility rules
    visible_user_ids = get_visible_task_user_ids(conn, user_id)
    
    # Show tasks assigned to visible users (current user's tasks + Alex Carter's tasks, 
    # or all tasks if current user is Alex Carter)
    if visible_user_ids:
        placeholders = ','.join(['%s'] * len(visible_user_ids))
        cur.execute(
            f"""SELECT t.id, t.title, t.status, t.priority, t.due_date, t.closure_date, 
                       COALESCE(t.work_type, 'Task') AS work_type,
                       t.assigned_to, t.project_id, p.name AS project_name
               FROM tasks t
               LEFT JOIN projects p ON p.id = t.project_id
               WHERE t.assigned_type='member' AND t.assigned_to IN ({placeholders})
               ORDER BY FIELD(t.status,'Open','In Progress','Review','Blocked','Closed'),
                        t.due_date IS NULL, t.due_date ASC""",
            tuple(visible_user_ids),
        )
        tasks = cur.fetchall()
    else:
        tasks = []
    
    cur.close()

    today = datetime.date.today()
    return render(request, "core/tasks_my.html", {"tasks": tasks, "page": "my_tasks", "today": today, "current_user_id": user_id})


# ==============================
#  UNASSIGNED TASKS
# ==============================
def unassigned_tasks_view(request):
    conn = get_tenant_conn(request)
    cur = conn.cursor()

    sql = """
        SELECT 
            t.id, 
            t.title,
            t.description,
            t.status,
            t.priority,
            t.due_date,
            t.created_at,
            t.assigned_type, 
            t.assigned_to,
            t.project_id,
            p.name AS project_name,
            sp.name AS subproject_name
        FROM tasks t
        LEFT JOIN projects p ON p.id = t.project_id
        LEFT JOIN subprojects sp ON sp.id = t.subproject_id
        WHERE (t.assigned_to IS NULL OR t.assigned_to = '')
          AND t.status NOT IN ('Blocked', 'Closed')
          AND t.status IN ('Open', 'In Progress', 'Review', 'Pending', 'New')
        ORDER BY 
            CASE WHEN t.due_date IS NULL THEN 1 ELSE 0 END,
            t.due_date ASC,
            t.created_at ASC                            
            """    
    cur.execute(sql)
    rows = cur.fetchall()

    # also needed for modal dropdown
    cur.execute("SELECT id, email, first_name, last_name FROM members ORDER BY first_name")
    members = cur.fetchall()
    cur.execute("SELECT id, name FROM teams ORDER BY name")
    teams = cur.fetchall()
    cur.close()

    return render(
        request,
        "core/unassigned_tasks.html",
        {"tasks": rows, "members": members, "teams": teams, "page": "unassigned_tasks"},
    )


# ==============================
#  TASK BOARD PAGE + DATA
# ==============================

from django.http import JsonResponse
from math import ceil
from .db_helpers import get_tenant_conn


def board_data_api(request):
    """
    Paginated task data for the Kanban board.
    Compatible with both dict and tuple-returning cursors.
    """
    conn = get_tenant_conn(request)
    cur = conn.cursor()

    # Get current user ID
    user_id = request.session.get("user_id")
    if not user_id:
        user_id = request.session.get("member_id")
    
    # Get visible task user IDs based on Alex Carter visibility rules
    visible_user_ids = get_visible_task_user_ids(conn, user_id) if user_id else []

    # ---- Status Filter + Assigned Filter ----
    status = request.GET.get("status")
    assigned_to = request.GET.get("assigned_to")
    exclude_closed = request.GET.get("exclude_closed")  # For timer dropdown

    # ---- Pagination ----
    try:
        page = int(request.GET.get("page", "1"))
    except:
        page = 1
    if page < 1:
        page = 1

    per_page = 20
    offset = (page - 1) * per_page

    # ---- Total Count Query ----
    count_sql = "SELECT COUNT(*) FROM tasks"
    count_params = []

    # build optional filters
    count_filters = []
    
    # Apply visibility filter for member-assigned tasks
    if visible_user_ids:
        placeholders = ','.join(['%s'] * len(visible_user_ids))
        count_filters.append(f"(assigned_type = 'member' AND assigned_to IN ({placeholders}))")
        count_params.extend(visible_user_ids)
    
    if status:
        count_filters.append("status = %s")
        count_params.append(status)
    if assigned_to:
        # Only member-assigned tasks should be considered when filtering by assigned_to
        count_filters.append("assigned_type = 'member' AND assigned_to = %s")
        count_params.append(assigned_to)
    if exclude_closed:
        # Exclude closed, cancelled, and completed tasks (for timer dropdown)
        count_filters.append("status NOT IN ('Closed', 'Cancelled', 'Completed')")

    if count_filters:
        count_sql += " WHERE " + " AND ".join(count_filters)

    cur.execute(count_sql, tuple(count_params))
    row = cur.fetchone()

    # Handle dict / tuple cursor
    if isinstance(row, dict):
        total_count = row.get("COUNT(*)", 0)
    else:
        total_count = row[0]

    # Calculate total pages
    total_pages = (total_count + per_page - 1) // per_page if total_count > 0 else 1

    # ---- MAIN QUERY ----
    sql = """
        SELECT
            t.id,
            COALESCE(t.title, '(Untitled)') AS title,
            COALESCE(t.status, 'Open') AS status,
            COALESCE(t.priority, 'Normal') AS priority,
            t.due_date,
            COALESCE(t.work_type, 'Task') AS work_type,

            CONCAT_WS(':', t.assigned_type, t.assigned_to) AS assigned_to,

            CASE
                WHEN t.assigned_type = 'member' THEN (
                    SELECT CONCAT(m.first_name, ' ', m.last_name)
                    FROM members m
                    WHERE m.id = t.assigned_to
                )
                WHEN t.assigned_type = 'team' THEN (
                    SELECT tm.name
                    FROM teams tm
                    WHERE tm.id = t.assigned_to
                )
                ELSE NULL
            END AS assigned_to_display

        FROM tasks t
    """

    # ---- Apply Status Filter ----
    # ---- Apply optional filters to main query ----
    params = []
    main_filters = []
    
    # Apply visibility filter for member-assigned tasks
    if visible_user_ids:
        placeholders = ','.join(['%s'] * len(visible_user_ids))
        main_filters.append(f"(t.assigned_type = 'member' AND t.assigned_to IN ({placeholders}))")
        params.extend(visible_user_ids)
    
    if status:
        main_filters.append("t.status = %s")
        params.append(status)
    if assigned_to:
        main_filters.append("t.assigned_type = 'member' AND t.assigned_to = %s")
        params.append(assigned_to)
    if exclude_closed:
        # Exclude closed, cancelled, and completed tasks (for timer dropdown)
        main_filters.append("t.status NOT IN ('Closed', 'Cancelled', 'Completed')")

    if main_filters:
        sql += " WHERE " + " AND ".join(main_filters)

    # ---- ORDER + Pagination ----
    sql += " ORDER BY t.id DESC LIMIT %s OFFSET %s"
    params.extend([per_page, offset])

    cur.execute(sql, tuple(params))

    # ---- Convert rows ----
    rows = cur.fetchall()

    if not rows:
        tasks = []
    elif isinstance(rows[0], dict):
        tasks = rows
    else:
        cols = [desc[0] for desc in cur.description]
        tasks = [dict(zip(cols, r)) for r in rows]

    cur.close()

    return JsonResponse({
        "tasks": tasks,
        "page": page,
        "per_page": per_page,
        "total_count": total_count,
        "total_pages": total_pages
    })





# ==============================
#  ASSIGN TASK (AJAX)
# ==============================
@require_POST
def assign_task_api(request):
    conn = get_tenant_conn(request)
    cur = conn.cursor()
    task_id = request.POST.get("task_id")
    assignee = request.POST.get("assignee")  # "member:23" or "team:5"
    assigned_by = request.session.get("user_id")

    if not task_id or not assignee:
        return HttpResponseBadRequest("Missing parameters")

    assigned_type, assigned_to = None, None
    if ":" in assignee:
        assigned_type, assigned_to = assignee.split(":", 1)
    else:
        assigned_type, assigned_to = "member", assignee

    # Get task details for notification
    cur.execute("SELECT title FROM tasks WHERE id=%s", (task_id,))
    task = cur.fetchone()
    task_title = task['title'] if task else 'A task'
    
    cur.execute(
        "UPDATE tasks SET assigned_to=%s, assigned_type=%s, updated_at=NOW() WHERE id=%s",
        (assigned_to, assigned_type, task_id),
    )
    cur.execute(
        "INSERT INTO activity_log (entity_type, entity_id, action, performed_by) VALUES (%s,%s,%s,%s)",
        ("task", task_id, f"assigned_to:{assignee}", assigned_by),
    )
    
    # Create notification for assigned member/team
    if assigned_type == "member":
        # Get assigner name
        cur.execute("SELECT CONCAT(first_name, ' ', last_name) as name FROM members WHERE id=%s", (assigned_by,))
        assigner = cur.fetchone()
        assigner_name = assigner['name'] if assigner else 'Someone'
        
        cur.execute("""
            INSERT INTO notifications (user_id, title, message, type, link)
            VALUES (%s, %s, %s, %s, %s)
        """, (
            assigned_to,
            "Task Assigned to You",
            f"{assigner_name} assigned you to '{task_title}'",
            "task",
            f"/tasks/{task_id}/view/"
        ))
    
    elif assigned_type == "team":
        cur.execute("SELECT CONCAT(first_name, ' ', last_name) as name FROM members WHERE id=%s", (assigned_by,))
        assigner = cur.fetchone()
        assigner_name = assigner['name'] if assigner else 'Someone'
        
        # Get all team members
        cur.execute("SELECT member_id FROM team_memberships WHERE team_id=%s", (assigned_to,))
        team_members = cur.fetchall()
        
        for member in team_members:
            cur.execute("""
                INSERT INTO notifications (user_id, title, message, type, link)
                VALUES (%s, %s, %s, %s, %s)
            """, (
                member['member_id'],
                "Team Task Assignment",
                f"{assigner_name} assigned a task to your team: '{task_title}'",
                "task",
                f"/tasks/{task_id}/view/"
            ))
    
    conn.commit()
    cur.close()
    return JsonResponse({"ok": True, "task_id": task_id, "assignee": assignee})


# ==============================
#  UPDATE STATUS (AJAX)
# ==============================
@require_POST
def api_update_status(request):
    """Called from Kanban drag-drop to update status"""

    conn = get_tenant_conn(request)
    cur = conn.cursor()

    task_id = request.POST.get("task_id")
    new_status = request.POST.get("status")
    user_id = request.session.get("user_id")

    # 1. CHECK REQUIRED PARAMS
    if not task_id or not new_status:
        return HttpResponseBadRequest("Missing parameters")

    # 2. SET CLOSURE DATE ONLY WHEN STATUS BECOMES 'Closed'
    if new_status == "Closed":
        cur.execute("""
            UPDATE tasks
            SET status=%s,
                closure_date=NOW(),
                updated_at=NOW()
            WHERE id=%s
        """, (new_status, task_id))
    else:
        # 3. FOR ANY OTHER STATUS → remove closure date?
        cur.execute("""
            UPDATE tasks
            SET status=%s,
                closure_date=NULL,
                updated_at=NOW()
            WHERE id=%s
        """, (new_status, task_id))

    # 4. CREATE LOG ENTRY
    cur.execute("""
        INSERT INTO activity_log (entity_type, entity_id, action, performed_by)
        VALUES (%s, %s, %s, %s)
    """, ("task", task_id, f"status_changed:{new_status}", user_id))

    # 5. CREATE NOTIFICATION WHEN TASK IS COMPLETED
    if new_status == "Closed" or new_status == "Completed":
        # Get task details and creator
        cur.execute("""
            SELECT t.title, t.created_by, t.assigned_to, t.assigned_type,
                   CONCAT(m.first_name, ' ', m.last_name) as updater_name
            FROM tasks t
            LEFT JOIN members m ON m.id = %s
            WHERE t.id = %s
        """, (user_id, task_id))
        task_info = cur.fetchone()
        
        if task_info:
            updater_name = task_info['updater_name'] or 'Someone'
            task_title = task_info['title'] or 'A task'
            
            # Notify task creator if they're not the one who completed it
            if task_info['created_by'] and str(task_info['created_by']) != str(user_id):
                cur.execute("""
                    INSERT INTO notifications (user_id, title, message, type, link)
                    VALUES (%s, %s, %s, %s, %s)
                """, (
                    task_info['created_by'],
                    "Task Completed",
                    f"{updater_name} completed task '{task_title}'",
                    "success",
                    f"/tasks/{task_id}/view/"
                ))
            
            # Notify assigned member if they're not the one who completed it
            if task_info['assigned_type'] == 'member' and task_info['assigned_to'] and str(task_info['assigned_to']) != str(user_id):
                cur.execute("""
                    INSERT INTO notifications (user_id, title, message, type, link)
                    VALUES (%s, %s, %s, %s, %s)
                """, (
                    task_info['assigned_to'],
                    "Task Completed",
                    f"{updater_name} marked '{task_title}' as complete",
                    "success",
                    f"/tasks/{task_id}/view/"
                ))

    # 6. SAVE CHANGES
    conn.commit()
    cur.close()

    return JsonResponse({"ok": True})


# ==============================
#  BULK IMPORT CSV
# ==============================
def download_excel_template(request):
    """Generate work-type-specific Excel template with dropdowns"""
    # Check authentication
    if not request.session.get('user_id'):
        return redirect('identify')
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
    from django.http import HttpResponse
    
    # Get work_type from query parameter (default to 'Task')
    work_type = request.GET.get('work_type', 'Task')
    
    # Define work-type-specific columns (work_type removed, file_attachment added)
    WORK_TYPE_COLUMNS = {
        'Task': ['title', 'description', 'project_name', 'subproject_name', 'priority', 
                 'due_date', 'assigned_to_name', 'file_attachment'],
        'Bug': ['title', 'project_name', 'subproject_name', 'severity', 'priority', 
                'status', 'due_date', 'assigned_to_name', 'description', 
                'steps_to_reproduce', 'expected_behavior', 'actual_behavior', 'file_attachment'],
        'Story': ['title', 'description', 'project_name', 'subproject_name', 'priority', 
                  'story_points', 'acceptance_criteria', 'due_date', 'assigned_to_name', 'file_attachment'],
        'Defect': ['title', 'project_name', 'subproject_name', 'severity', 'priority', 
                   'status', 'due_date', 'assigned_to_name', 'description', 
                   'steps_to_reproduce', 'expected_behavior', 'actual_behavior', 'file_attachment'],
        'Subtask': ['title', 'description', 'project_name', 'subproject_name', 'parent_task_id',
                    'priority', 'due_date', 'assigned_to_name', 'file_attachment'],
        'Report': ['title', 'description', 'project_name', 'subproject_name', 'report_type',
                   'priority', 'due_date', 'assigned_to_name', 'file_attachment'],
        'Change Request': ['title', 'description', 'project_name', 'subproject_name', 
                          'change_type', 'impact', 'priority', 'due_date', 'assigned_to_name', 'file_attachment']
    }
    
    # Get columns for selected work_type
    headers = WORK_TYPE_COLUMNS.get(work_type, WORK_TYPE_COLUMNS['Task'])
    
    conn = get_tenant_conn(request)
    cur = conn.cursor()
    
    try:
        # Fetch data for dropdowns
        cur.execute("SELECT id, name FROM projects ORDER BY name")
        projects = cur.fetchall()
        
        cur.execute("SELECT id, CONCAT(first_name, ' ', last_name) as name FROM members ORDER BY first_name")
        members = cur.fetchall()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Tasks Import"
        
        # Create hidden sheets for dropdown data
        projects_sheet = wb.create_sheet("Projects_Data")
        members_sheet = wb.create_sheet("Members_Data")
        subprojects_sheet = wb.create_sheet("Subprojects_Data")
        
        # Hide data sheets
        projects_sheet.sheet_state = 'hidden'
        members_sheet.sheet_state = 'hidden'
        subprojects_sheet.sheet_state = 'hidden'
        
        # Populate Projects data sheet
        projects_sheet['A1'] = 'project_id'
        projects_sheet['B1'] = 'project_name'
        for idx, proj in enumerate(projects, start=2):
            projects_sheet[f'A{idx}'] = proj['id']
            projects_sheet[f'B{idx}'] = proj['name']
        
        # Populate Members data sheet
        members_sheet['A1'] = 'member_id'
        members_sheet['B1'] = 'member_name'
        for idx, mem in enumerate(members, start=2):
            members_sheet[f'A{idx}'] = mem['id']
            members_sheet[f'B{idx}'] = mem['name']
        
        # Fetch all subprojects with project reference
        cur.execute("SELECT id, name, project_id FROM subprojects ORDER BY project_id, name")
        subprojects = cur.fetchall()
        
        # Populate Subprojects data sheet with project info for reference
        subprojects_sheet['A1'] = 'subproject_id'
        subprojects_sheet['B1'] = 'subproject_name'
        subprojects_sheet['C1'] = 'project_id'
        subprojects_sheet['D1'] = 'project_name'
        
        # Build project name lookup
        project_names = {proj['id']: proj['name'] for proj in projects}
        
        for idx, subproj in enumerate(subprojects, start=2):
            subprojects_sheet[f'A{idx}'] = subproj['id']
            subprojects_sheet[f'B{idx}'] = subproj['name']
            subprojects_sheet[f'C{idx}'] = subproj['project_id']
            subprojects_sheet[f'D{idx}'] = project_names.get(subproj['project_id'], '')
        
        # Setup main sheet headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Set column widths dynamically
        column_widths = {
            'title': 30, 'description': 40, 'project_name': 25, 'subproject_name': 25,
            'status': 12, 'priority': 12, 'severity': 12, 'assigned_to_name': 25,
            'due_date': 12, 'file_attachment': 40, 'steps_to_reproduce': 40,
            'expected_behavior': 35, 'actual_behavior': 35, 'story_points': 12,
            'acceptance_criteria': 40, 'parent_task_id': 15, 'report_type': 20,
            'change_type': 20, 'impact': 15
        }
        
        for col_num, header in enumerate(headers, start=1):
            col_letter = chr(64 + col_num)  # A=65, so 64+1=A
            ws.column_dimensions[col_letter].width = column_widths.get(header, 20)
        
        # Add data validations based on columns present
        col_index = {header: idx + 1 for idx, header in enumerate(headers)}
        
        # Project name validation
        if 'project_name' in col_index and len(projects) > 0:
            col_letter = chr(64 + col_index['project_name'])
            project_dv = DataValidation(
                type="list",
                formula1=f"=Projects_Data!$B$2:$B${len(projects)+1}",
                allow_blank=True
            )
            project_dv.error = 'Please select a valid project'
            project_dv.errorTitle = 'Invalid Project'
            ws.add_data_validation(project_dv)
            project_dv.add(f'{col_letter}2:{col_letter}1000')
        
        # Subproject name validation
        if 'subproject_name' in col_index and len(subprojects) > 0:
            col_letter = chr(64 + col_index['subproject_name'])
            subproject_dv = DataValidation(
                type="list",
                formula1=f"=Subprojects_Data!$B$2:$B${len(subprojects)+1}",
                allow_blank=True
            )
            subproject_dv.error = 'Please select a valid subproject for your project'
            subproject_dv.errorTitle = 'Invalid Subproject'
            ws.add_data_validation(subproject_dv)
            subproject_dv.add(f'{col_letter}2:{col_letter}1000')
        
        # Status validation
        if 'status' in col_index:
            col_letter = chr(64 + col_index['status'])
            status_dv = DataValidation(
                type="list",
                formula1='"Open,In Progress,Review,Blocked,Closed,Pending,New"',
                allow_blank=True
            )
            status_dv.error = 'Please select a valid status'
            status_dv.errorTitle = 'Invalid Status'
            ws.add_data_validation(status_dv)
            status_dv.add(f'{col_letter}2:{col_letter}1000')
        
        # Priority validation
        if 'priority' in col_index:
            col_letter = chr(64 + col_index['priority'])
            priority_dv = DataValidation(
                type="list",
                formula1='"Low,Normal,High,Critical"',
                allow_blank=True
            )
            priority_dv.error = 'Please select a valid priority'
            priority_dv.errorTitle = 'Invalid Priority'
            ws.add_data_validation(priority_dv)
            priority_dv.add(f'{col_letter}2:{col_letter}1000')
        
        # Severity validation (for Bug/Defect)
        if 'severity' in col_index:
            col_letter = chr(64 + col_index['severity'])
            severity_dv = DataValidation(
                type="list",
                formula1='"Low,Medium,High,Critical"',
                allow_blank=True
            )
            severity_dv.error = 'Please select a valid severity'
            severity_dv.errorTitle = 'Invalid Severity'
            ws.add_data_validation(severity_dv)
            severity_dv.add(f'{col_letter}2:{col_letter}1000')
        
        # Assigned to validation
        if 'assigned_to_name' in col_index and len(members) > 0:
            col_letter = chr(64 + col_index['assigned_to_name'])
            member_dv = DataValidation(
                type="list",
                formula1=f"=Members_Data!$B$2:$B${len(members)+1}",
                allow_blank=True
            )
            member_dv.error = 'Please select a valid member'
            member_dv.errorTitle = 'Invalid Member'
            ws.add_data_validation(member_dv)
            member_dv.add(f'{col_letter}2:{col_letter}1000')
        
        # Note: work_type column removed - it's determined by the template downloaded
        # file_attachment column is for file paths (no validation needed)
        
        # Add sample row based on work_type
        sample_data = {
            'Task': {
                'title': 'Fix login bug',
                'description': 'Users cannot login with special characters in password',
                'priority': 'High',
                'due_date': '2026-03-15',
                'file_attachment': 'C:\\path\\to\\screenshot.png'
            },
            'Bug': {
                'title': 'Login fails with special characters',
                'description': 'Users cannot login when password contains special characters',
                'severity': 'High',
                'priority': 'Critical',
                'status': 'Open',
                'due_date': '2026-03-15',
                'steps_to_reproduce': '1. Enter username\n2. Enter password with @ symbol\n3. Click login',
                'expected_behavior': 'User should be logged in successfully',
                'actual_behavior': 'Error message displayed: Invalid credentials',
                'file_attachment': 'C:\\path\\to\\error_screenshot.png'
            },
            'Story': {
                'title': 'User can reset password',
                'description': 'As a user, I want to reset my password so that I can regain access to my account',
                'priority': 'High',
                'story_points': '5',
                'acceptance_criteria': '- User receives reset email\n- Link expires after 24 hours\n- Password meets security requirements',
                'due_date': '2026-03-20',
                'file_attachment': 'C:\\path\\to\\mockup.png'
            },
            'Defect': {
                'title': 'Dashboard shows incorrect data',
                'description': 'Dashboard displays wrong task counts',
                'severity': 'Medium',
                'priority': 'High',
                'status': 'New',
                'due_date': '2026-03-18',
                'steps_to_reproduce': '1. Login to dashboard\n2. Check task count widget',
                'expected_behavior': 'Shows correct count of tasks',
                'actual_behavior': 'Shows count from previous day',
                'file_attachment': 'C:\\path\\to\\dashboard_screenshot.png'
            },
            'Subtask': {
                'title': 'Design login form mockup',
                'description': 'Create mockup for new login form design',
                'parent_task_id': '123',
                'priority': 'Normal',
                'due_date': '2026-03-12',
                'file_attachment': 'C:\\path\\to\\design.png'
            },
            'Report': {
                'title': 'Monthly performance report',
                'description': 'Generate monthly performance metrics report',
                'report_type': 'Performance',
                'priority': 'Normal',
                'due_date': '2026-03-31',
                'file_attachment': 'C:\\path\\to\\report_data.xlsx'
            },
            'Change Request': {
                'title': 'Add two-factor authentication',
                'description': 'Implement 2FA for enhanced security',
                'change_type': 'Enhancement',
                'impact': 'High',
                'priority': 'High',
                'due_date': '2026-04-15',
                'file_attachment': 'C:\\path\\to\\requirements.pdf'
            }
        }
        
        sample = sample_data.get(work_type, sample_data['Task'])
        
        # Fill sample row
        for col_num, header in enumerate(headers, start=1):
            col_letter = chr(64 + col_num)
            if header == 'project_name':
                ws[f'{col_letter}2'] = projects[0]['name'] if projects else ''
            elif header == 'subproject_name':
                # Find a subproject for the first project
                first_project_subproject = None
                if projects and subprojects:
                    for subproj in subprojects:
                        if subproj['project_id'] == projects[0]['id']:
                            first_project_subproject = subproj['name']
                            break
                ws[f'{col_letter}2'] = first_project_subproject if first_project_subproject else ''
            elif header == 'assigned_to_name':
                ws[f'{col_letter}2'] = members[0]['name'] if members else ''
            elif header in sample:
                ws[f'{col_letter}2'] = sample[header]
            else:
                ws[f'{col_letter}2'] = ''
        
        # Create response with work-type-specific filename
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'tasks_import_{work_type.lower().replace(" ", "_")}_template.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        wb.save(response)
        return response
        
    except Exception as e:
        return HttpResponse(f"Error generating template: {str(e)}", status=500)
    finally:
        cur.close()


def download_csv_template(request):
    """Generate CSV template with reference data in comments"""
    # Check authentication
    if not request.session.get('user_id'):
        return redirect('identify')
    
    from django.http import HttpResponse
    import csv

    conn = get_tenant_conn(request)
    cur = conn.cursor()

    try:
        # Fetch data for reference
        cur.execute("SELECT id, name FROM projects ORDER BY name")
        projects = cur.fetchall()

        cur.execute("SELECT id, CONCAT(first_name, ' ', last_name) as name FROM members ORDER BY first_name")
        members = cur.fetchall()

        cur.execute("SELECT id, name, project_id FROM subprojects ORDER BY project_id, name")
        subprojects = cur.fetchall()

        # Create response
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=tasks_import_template.csv'

        writer = csv.writer(response)

        # Write reference data as comments
        writer.writerow(['# BULK IMPORT TEMPLATE - REFERENCE DATA'])
        writer.writerow(['# '])
        writer.writerow(['# AVAILABLE PROJECTS:'])
        for proj in projects:
            writer.writerow([f'#   project_id: {proj["id"]} - {proj["name"]}'])

        writer.writerow(['# '])
        writer.writerow(['# AVAILABLE MEMBERS:'])
        for mem in members:
            writer.writerow([f'#   member_id: {mem["id"]} - {mem["name"]}'])

        writer.writerow(['# '])
        writer.writerow(['# AVAILABLE SUBPROJECTS:'])
        for subproj in subprojects:
            writer.writerow([f'#   subproject_id: {subproj["id"]} - {subproj["name"]} (project: {subproj["project_id"]})'])

        writer.writerow(['# '])
        writer.writerow(['# VALID STATUS: Open, In Progress, Review, Blocked, Closed, Pending, New'])
        writer.writerow(['# VALID PRIORITY: Low, Normal, High, Critical'])
        writer.writerow(['# VALID WORK_TYPE: Task, Bug, Story, Defect, Subtask, Report, Change Request'])
        writer.writerow(['# DATE FORMAT: YYYY-MM-DD (e.g., 2026-03-15)'])
        writer.writerow(['# ASSIGNED_TO FORMAT: member:ID or team:ID (e.g., member:5 or team:2)'])
        writer.writerow(['# '])
        writer.writerow(['# DELETE ALL COMMENT LINES (starting with #) BEFORE UPLOADING'])
        writer.writerow(['# '])

        # Write header row
        writer.writerow(['title', 'description', 'project_id', 'subproject_id', 'status',
                        'priority', 'assigned_to', 'due_date', 'work_type'])

        # Write sample row
        if projects:
            writer.writerow([
                'Fix login bug',
                'Users cannot login with special characters in password',
                projects[0]['id'],
                '',
                'Open',
                'High',
                f'member:{members[0]["id"]}' if members else '',
                '2026-03-15',
                'Task'
            ])

        return response

    except Exception as e:
        return HttpResponse(f"Error generating template: {str(e)}", status=500)
    finally:
        cur.close()



def bulk_import_csv_view(request):
    """Upload & import CSV file of tasks with comprehensive validation"""
    context = {"page": "bulk_import"}

    if request.method == "POST" and request.FILES.get("csv_file"):
        # Get work_type from POST data (passed from the form)
        work_type = request.POST.get("work_type", "Task")
        
        conn = get_tenant_conn(request)
        cur = conn.cursor()
        
        try:
            f = request.FILES["csv_file"]
            
            # Validate file size (max 5MB)
            if f.size > 5 * 1024 * 1024:
                context.update({
                    "inserted": 0, 
                    "errors": [{"row": 0, "error": "File size exceeds 5MB limit", "data": {}}]
                })
                cur.close()
                return render(request, "core/tasks_bulk_import.html", context)
            
            # Validate file extension - accept both CSV and Excel
            file_ext = f.name.lower()
            if not (file_ext.endswith('.csv') or file_ext.endswith('.xlsx') or file_ext.endswith('.xls')):
                context.update({
                    "inserted": 0, 
                    "errors": [{"row": 0, "error": "Invalid file type. Please upload a CSV or Excel file (.csv, .xlsx, .xls)", "data": {}}]
                })
                cur.close()
                return render(request, "core/tasks_bulk_import.html", context)
            
            # Read file based on type
            if file_ext.endswith('.csv'):
                # Read and decode CSV file
                try:
                    text = f.read().decode("utf-8")
                except UnicodeDecodeError:
                    try:
                        f.seek(0)
                        text = f.read().decode("utf-8-sig")  # Try with BOM
                    except Exception:
                        context.update({
                            "inserted": 0, 
                            "errors": [{"row": 0, "error": "Unable to decode CSV file. Please ensure it's UTF-8 encoded.", "data": {}}]
                        })
                        cur.close()
                        return render(request, "core/tasks_bulk_import.html", context)
                
                reader = csv.DictReader(io.StringIO(text))
            else:
                # Read Excel file
                from openpyxl import load_workbook
                try:
                    wb = load_workbook(f, data_only=True)
                    ws = wb.active
                    
                    # Convert Excel to dict format
                    data = []
                    headers = []
                    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                        if row_idx == 1:
                            headers = [str(cell).strip() if cell else '' for cell in row]
                        else:
                            if any(row):  # Skip empty rows
                                row_dict = {}
                                for col_idx, value in enumerate(row):
                                    if col_idx < len(headers):
                                        # Convert None to empty string, handle dates
                                        if value is None:
                                            row_dict[headers[col_idx]] = ''
                                        elif hasattr(value, 'strftime'):  # Date object
                                            row_dict[headers[col_idx]] = value.strftime('%Y-%m-%d')
                                        else:
                                            row_dict[headers[col_idx]] = str(value).strip()
                                data.append(row_dict)
                    
                    # Create a simple class to mimic csv.DictReader
                    class ExcelReader:
                        def __init__(self, data, fieldnames):
                            self.data = data
                            self.fieldnames = fieldnames
                        
                        def __iter__(self):
                            return iter(self.data)
                    
                    reader = ExcelReader(data, headers)
                    
                except Exception as e:
                    context.update({
                        "inserted": 0, 
                        "errors": [{"row": 0, "error": f"Unable to read Excel file: {str(e)}", "data": {}}]
                    })
                    cur.close()
                    return render(request, "core/tasks_bulk_import.html", context)
            
            inserted, errors, warnings = 0, [], []

            # Enforce required header columns - Support both old (ID) and new (name) formats
            required_cols_old = {"title", "project_id"}
            required_cols_new = {"title", "project_name"}
            
            # Extended optional columns to support work-type-specific fields
            optional_cols_old = {"description", "subproject_id", "status", "priority", "assigned_to", "due_date", "work_type", "file_attachment"}
            optional_cols_new = {
                "description", "subproject_name", "status", "priority", "assigned_to_name", "due_date", "work_type", "note", "file_attachment",
                # Bug/Defect specific
                "severity", "steps_to_reproduce", "expected_behavior", "actual_behavior",
                # Story specific
                "story_points", "acceptance_criteria",
                # Subtask specific
                "parent_task_id",
                # Report specific
                "report_type",
                # Change Request specific
                "change_type", "impact"
            }
            
            present = set((reader.fieldnames or []))
            
            # Check if using old format (IDs) or new format (names)
            using_names = "project_name" in present
            
            if using_names:
                required_cols = required_cols_new
                optional_cols = optional_cols_new
            else:
                required_cols = required_cols_old
                optional_cols = optional_cols_old
            
            missing = required_cols - present
            if missing:
                errors.append({
                    "row": 0, 
                    "error": f"Missing required columns: {', '.join(sorted(missing))}", 
                    "data": {}
                })
                context.update({"inserted": 0, "errors": errors})
                cur.close()
                return render(request, "core/tasks_bulk_import.html", context)
            
            # Warn about unexpected columns
            expected = required_cols | optional_cols
            unexpected = present - expected
            if unexpected:
                warnings.append(f"Unexpected columns will be ignored: {', '.join(sorted(unexpected))}")

            # Build lookup dictionaries for name-to-ID conversion
            cur.execute("SELECT id, name FROM projects")
            projects_dict = {row['name']: row['id'] for row in cur.fetchall()}
            projects_dict_by_id = {str(row['id']): row['id'] for row in cur.fetchall()}
            
            cur.execute("SELECT id FROM projects")
            valid_projects = {str(row['id']) for row in cur.fetchall()}
            
            cur.execute("SELECT id, CONCAT(first_name, ' ', last_name) as name FROM members")
            members_dict = {row['name']: row['id'] for row in cur.fetchall()}
            
            cur.execute("SELECT id FROM members")
            valid_members = {str(row['id']) for row in cur.fetchall()}
            
            cur.execute("SELECT id, name, project_id FROM subprojects")
            subprojects_list = cur.fetchall()
            subprojects_dict = {(row['name'], row['project_id']): row['id'] for row in subprojects_list}
            
            cur.execute("SELECT id FROM teams")
            valid_teams = {str(row['id']) for row in cur.fetchall()}

            # Process each row
            for i, raw_row in enumerate(reader, start=1):
                try:
                    # Normalize/strip all values
                    row = {k: (v.strip() if isinstance(v, str) else v) for k, v in raw_row.items()}

                    # Validate required fields
                    if not row.get("title"):
                        raise Exception("Missing required field: title")
                    
                    # Handle project - convert name to ID if using new format
                    if using_names:
                        project_name = row.get("project_name", "").strip()
                        if not project_name:
                            raise Exception("Missing required field: project_name")
                        
                        if project_name not in projects_dict:
                            raise Exception(f"Invalid project_name: '{project_name}'. Project does not exist.")
                        
                        project_id = projects_dict[project_name]
                    else:
                        if not row.get("project_id"):
                            raise Exception("Missing required field: project_id")
                        
                        project_id = str(row.get("project_id")).strip()
                        if project_id not in valid_projects:
                            raise Exception(f"Invalid project_id: {project_id}. Project does not exist.")
                        project_id = int(project_id)

                    # Handle subproject - convert name to ID if using new format
                    subproject_id = None
                    if using_names:
                        subproject_name = row.get("subproject_name", "").strip()
                        if subproject_name:
                            # Look up subproject by name and project_id
                            subproject_key = (subproject_name, project_id)
                            if subproject_key not in subprojects_dict:
                                raise Exception(f"Invalid subproject_name: '{subproject_name}' for selected project")
                            subproject_id = subprojects_dict[subproject_key]
                    else:
                        subproject_id_str = row.get("subproject_id", "").strip()
                        if subproject_id_str:
                            cur.execute("SELECT id FROM subprojects WHERE id=%s AND project_id=%s", 
                                      (subproject_id_str, project_id))
                            if not cur.fetchone():
                                raise Exception(f"Invalid subproject_id: {subproject_id_str} for project {project_id}")
                            subproject_id = int(subproject_id_str)

                    # Parse assigned_to - handle both old format (IDs) and new format (names)
                    assigned_to, assigned_type = None, None
                    
                    if using_names:
                        # New format: assigned_to_name column with member names
                        assigned_name = row.get("assigned_to_name", "").strip()
                        if assigned_name:
                            if assigned_name not in members_dict:
                                raise Exception(f"Invalid assigned_to_name: '{assigned_name}'. Member does not exist.")
                            assigned_to = members_dict[assigned_name]
                            assigned_type = "member"
                    else:
                        # Old format: assigned_to column with 'member:ID' or 'team:ID' or plain ID
                        assigned_raw = row.get("assigned_to", "").strip()
                        if assigned_raw:
                            if ":" in assigned_raw:
                                atype, aid = assigned_raw.split(":", 1)
                                atype = atype.strip().lower()
                                aid = aid.strip()
                                if atype not in ("member", "team"):
                                    raise Exception(f"Invalid assigned_to type '{atype}'. Use 'member:ID' or 'team:ID'.")
                                if not aid.isdigit():
                                    raise Exception(f"Invalid assigned_to id '{aid}'. Must be numeric.")
                                
                                # Validate member/team exists
                                if atype == "member" and aid not in valid_members:
                                    raise Exception(f"Invalid member ID: {aid}. Member does not exist.")
                                elif atype == "team" and aid not in valid_teams:
                                    raise Exception(f"Invalid team ID: {aid}. Team does not exist.")
                                
                                assigned_type, assigned_to = atype, int(aid)
                            else:
                                # Assume member id
                                aid = assigned_raw.strip()
                                if aid and not aid.isdigit():
                                    raise Exception(f"Invalid assigned_to value '{aid}'. Use 'member:ID', 'team:ID', or numeric member ID.")
                                if aid and aid not in valid_members:
                                    raise Exception(f"Invalid member ID: {aid}. Member does not exist.")
                                assigned_type, assigned_to = "member", int(aid) if aid else None

                    # Validate status
                    status = row.get("status") or "Open"
                    valid_statuses = ["Open", "In Progress", "Review", "Blocked", "Closed", "Pending", "New"]
                    if status not in valid_statuses:
                        raise Exception(f"Invalid status: {status}. Must be one of: {', '.join(valid_statuses)}")

                    # Validate priority
                    priority = row.get("priority") or "Normal"
                    valid_priorities = ["Low", "Normal", "High", "Critical"]
                    if priority not in valid_priorities:
                        raise Exception(f"Invalid priority: {priority}. Must be one of: {', '.join(valid_priorities)}")

                    # Validate due_date (empty allowed)
                    due_date = row.get("due_date") or None
                    if due_date:
                        try:
                            # Accept YYYY-MM-DD only
                            datetime.datetime.strptime(due_date, "%Y-%m-%d")
                        except Exception:
                            raise Exception("Invalid due_date. Use YYYY-MM-DD format (e.g., 2024-12-31).")

                    # Optional work_type - use from Excel if present, otherwise use from form
                    row_work_type = row.get("work_type", "").strip()
                    if not row_work_type:
                        row_work_type = work_type  # Use work_type from form/URL
                    
                    # Build description with work-type-specific fields
                    base_description = row.get("description") or ""
                    full_description = base_description
                    
                    # Add work-type-specific fields to description
                    if row_work_type in ["Bug", "Defect"]:
                        # Validate severity for bugs/defects
                        severity = row.get("severity", "")
                        if severity:
                            valid_severities = ["Low", "Medium", "High", "Critical"]
                            if severity not in valid_severities:
                                raise Exception(f"Invalid severity: {severity}. Must be one of: {', '.join(valid_severities)}")
                            # Use severity as priority for bugs/defects
                            priority = severity
                        
                        # Append bug-specific fields to description
                        if row.get("steps_to_reproduce"):
                            full_description += f"\n\n**Steps to Reproduce:**\n{row.get('steps_to_reproduce')}"
                        if row.get("expected_behavior"):
                            full_description += f"\n\n**Expected Behavior:**\n{row.get('expected_behavior')}"
                        if row.get("actual_behavior"):
                            full_description += f"\n\n**Actual Behavior:**\n{row.get('actual_behavior')}"
                    
                    elif row_work_type == "Story":
                        # Append story-specific fields to description
                        if row.get("story_points"):
                            full_description += f"\n\n**Story Points:** {row.get('story_points')}"
                        if row.get("acceptance_criteria"):
                            full_description += f"\n\n**Acceptance Criteria:**\n{row.get('acceptance_criteria')}"
                    
                    elif row_work_type == "Subtask":
                        # Handle parent_task_id
                        parent_task_id = row.get("parent_task_id", "").strip()
                        if parent_task_id:
                            # Validate parent task exists
                            cur.execute("SELECT id FROM tasks WHERE id=%s", (parent_task_id,))
                            if not cur.fetchone():
                                raise Exception(f"Invalid parent_task_id: {parent_task_id}. Parent task does not exist.")
                            full_description += f"\n\n**Parent Task ID:** {parent_task_id}"
                    
                    elif row_work_type == "Report":
                        # Append report-specific fields
                        if row.get("report_type"):
                            full_description += f"\n\n**Report Type:** {row.get('report_type')}"
                    
                    elif row_work_type == "Change Request":
                        # Append change request-specific fields
                        if row.get("change_type"):
                            full_description += f"\n\n**Change Type:** {row.get('change_type')}"
                        if row.get("impact"):
                            full_description += f"\n\n**Impact:** {row.get('impact')}"

                    # Insert task
                    cur.execute(
                        """INSERT INTO tasks
                           (project_id, subproject_id, title, description, status, priority, work_type,
                            assigned_to, assigned_type, created_by, due_date, created_at)
                           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())""",
                        (
                            int(project_id),
                            int(subproject_id) if subproject_id else None,
                            row["title"],
                            full_description,  # Use full_description with work-type-specific fields
                            status,
                            priority,
                            row_work_type,  # Use row_work_type
                            assigned_to,
                            assigned_type,
                            request.session.get("user_id"),
                            due_date,
                        ),
                    )
                    
                    task_id = cur.lastrowid
                    
                    # Handle file attachment if provided
                    file_path = row.get("file_attachment", "").strip()
                    if file_path:
                        import os
                        from django.core.files.base import ContentFile
                        
                        # Check if file exists
                        if os.path.exists(file_path):
                            try:
                                # Read the file
                                with open(file_path, 'rb') as f:
                                    file_content = f.read()
                                
                                # Get filename
                                filename = os.path.basename(file_path)
                                
                                # Generate unique filename
                                import time
                                timestamp = int(time.time())
                                unique_filename = f"{task_id}_{timestamp}_{filename}"
                                
                                # Save to media/task_attachments/
                                from django.conf import settings
                                upload_dir = os.path.join(settings.MEDIA_ROOT, 'task_attachments')
                                os.makedirs(upload_dir, exist_ok=True)
                                
                                file_save_path = os.path.join(upload_dir, unique_filename)
                                with open(file_save_path, 'wb') as dest:
                                    dest.write(file_content)
                                
                                # Store relative path in database
                                relative_path = f'task_attachments/{unique_filename}'
                                
                                # Insert attachment record
                                cur.execute(
                                    """INSERT INTO task_attachments 
                                       (task_id, file_path, uploaded_by, uploaded_at) 
                                       VALUES (%s, %s, %s, NOW())""",
                                    (task_id, relative_path, request.session.get("user_id"))
                                )
                            except Exception as e:
                                warnings.append(f"Row {i}: Could not attach file '{file_path}': {str(e)}")
                        else:
                            warnings.append(f"Row {i}: File not found: '{file_path}'")
                    
                    # Log activity
                    try:
                        cur.execute(
                            "INSERT INTO activity_log (entity_type, entity_id, action, performed_by) VALUES (%s,%s,%s,%s)",
                            ("task", task_id, "created via bulk import", request.session.get("user_id")),
                        )
                    except Exception:
                        pass  # Don't fail import if logging fails
                    
                    inserted += 1
                    
                except Exception as e:
                    errors.append({"row": i, "error": str(e), "data": dict(raw_row)})

            conn.commit()
            context.update({
                "inserted": inserted, 
                "errors": errors,
                "warnings": warnings,
                "total_rows": i if 'i' in locals() else 0
            })
            
        except Exception as e:
            conn.rollback()
            context.update({
                "inserted": 0, 
                "errors": [{"row": 0, "error": f"Unexpected error: {str(e)}", "data": {}}]
            })
        finally:
            cur.close()
        
        return render(request, "core/tasks_bulk_import.html", context)

    return render(request, "core/tasks_bulk_import.html", context)


@require_GET
def api_task_detail(request):
    """
    GET ?id=<task_id>
    Returns JSON with task fields used by the modal: id, title, description, priority, due_date, assigned_to, assigned_to_display, status
    Implemented using tenant DB queries (no Django ORM model required).
    """
    tid = request.GET.get('id') or request.GET.get('task_id')
    if not tid:
        return JsonResponse({'error': 'missing id'}, status=400)

    conn = get_tenant_conn(request)
    cur = conn.cursor()
    try:
        # fetch task row and human-friendly assignee display name
        cur.execute("""
            SELECT
              t.id,
              COALESCE(t.title, '') AS title,
              COALESCE(t.description, '') AS description,
              COALESCE(t.priority, '') AS priority,
              t.due_date,
              t.status,
              t.project_id,
              t.subproject_id,
              t.assigned_to,
              t.assigned_type,
              CASE
                WHEN t.assigned_type = 'member' THEN (SELECT CONCAT(m.first_name, ' ', m.last_name) FROM members m WHERE m.id = t.assigned_to)
                WHEN t.assigned_type = 'team' THEN (SELECT tm.name FROM teams tm WHERE tm.id = t.assigned_to)
                ELSE NULL
              END AS assigned_to_display
            FROM tasks t
            WHERE t.id = %s
            LIMIT 1
        """, (tid,))
        row = cur.fetchone()
        if not row:
            return JsonResponse({'error': 'task not found'}, status=404)

        # map columns to dict depending on cursor return type
        if isinstance(row, dict):
            data = row
        else:
            cols = [desc[0] for desc in cur.description]
            data = dict(zip(cols, row))

        # ensure date is serialized as iso string (or empty)
        due = data.get('due_date')
        if due is None:
            data['due_date'] = ''
        else:
            # if it's a date/datetime object, convert to ISO date
            try:
                data['due_date'] = due.isoformat()
            except Exception:
                data['due_date'] = str(due)

        # normalize assigned fields to strings
        data['assigned_to'] = data.get('assigned_to') or ''
        data['assigned_to_display'] = data.get('assigned_to_display') or ''
        print("DATA:", data)
        return JsonResponse({
            'id': data.get('id'),
            'title': data.get('title', ''),
            'description': data.get('description', ''),
            'priority': data.get('priority', ''),
            'due_date': data.get('due_date', ''),
            'status': data.get('status', ''),
            'project_id': data.get('project_id'),
            'subproject_id': data.get('subproject_id'),
            'assigned_type': data.get('assigned_type'),
            'assigned_to': data.get('assigned_to', ''),
            'assigned_to_display': data.get('assigned_to_display', ''),
        })

    finally:
        cur.close()


@require_GET
def api_tasks_search(request):
    """
    Simple task search used by the global quick-search UI.
    GET params: q
    Returns: { tasks: [ { id, title, status, assigned_to_display, project_name } ] }
    """
    q = (request.GET.get('q') or '').strip()
    if not q:
        return JsonResponse({'tasks': []})

    conn = get_tenant_conn(request)
    cur = conn.cursor()
    try:
        like = f"%{q}%"
        cur.execute("""
            SELECT
                t.id,
                COALESCE(t.title, '') AS title,
                COALESCE(t.status, '') AS status,
                CASE
                    WHEN t.assigned_type = 'member' THEN (SELECT CONCAT(m.first_name, ' ', m.last_name) FROM members m WHERE m.id = t.assigned_to)
                    WHEN t.assigned_type = 'team' THEN (SELECT tm.name FROM teams tm WHERE tm.id = t.assigned_to)
                    ELSE NULL
                END AS assigned_to_display,
                (SELECT p.name FROM projects p WHERE p.id = t.project_id) AS project_name
            FROM tasks t
            WHERE (t.title LIKE %s OR t.description LIKE %s)
            ORDER BY t.id DESC
            LIMIT 20
        """, (like, like))

        rows = cur.fetchall() or []
        tasks = []
        if rows:
            if isinstance(rows[0], dict):
                for r in rows:
                    tasks.append({
                        'id': r.get('id'),
                        'title': r.get('title') or '',
                        'status': r.get('status') or '',
                        'assigned_to_display': r.get('assigned_to_display') or '',
                        'project_name': r.get('project_name') or ''
                    })
            else:
                cols = [d[0] for d in cur.description]
                for r in rows:
                    d = dict(zip(cols, r))
                    tasks.append({
                        'id': d.get('id'),
                        'title': d.get('title') or '',
                        'status': d.get('status') or '',
                        'assigned_to_display': d.get('assigned_to_display') or '',
                        'project_name': d.get('project_name') or ''
                    })

        return JsonResponse({'tasks': tasks})
    finally:
        cur.close()


@require_POST
def api_task_update(request):
    """
    Accepts form-data (task_id, title, description, priority, due_date, assigned_to_display, status)
    Returns { "ok": true } on success or { "ok": false, "error": "..." }
    Implemented using SQL updates against tenant DB (no ORM).
    """
    tid = request.POST.get('task_id')
    if not tid:
        return JsonResponse({'ok': False, 'error': 'missing task_id'}, status=400)

    conn = get_tenant_conn(request)
    cur = conn.cursor()
    try:
        # check exists
        cur.execute("SELECT id, title, status, created_by, assigned_to, assigned_type FROM tasks WHERE id = %s LIMIT 1", (tid,))
        existing = cur.fetchone()
        if not existing:
            return JsonResponse({'ok': False, 'error': 'task not found'}, status=404)

        # collect fields to update
        updates = []
        params = []

        title = request.POST.get('title')
        if title is not None:
            updates.append("title = %s")
            params.append(title)

        description = request.POST.get('description')
        if description is not None:
            updates.append("description = %s")
            params.append(description)

        priority = request.POST.get('priority')
        if priority is not None:
            updates.append("priority = %s")
            params.append(priority)

        due_date = request.POST.get('due_date')
        if due_date:
            # Accept YYYY-MM-DD (frontend uses input[type=date])
            updates.append("due_date = %s")
            params.append(due_date)
        elif due_date == '':
            # explicit empty string -> set NULL
            updates.append("due_date = NULL")

        # --- ASSIGNED TO (member:ID or team:ID) ---
        # frontend posts field name "assigned_to" (value like "member:23" or "team:5", or empty string to unassign)
        assigned_raw = request.POST.get('assigned_to')
        if assigned_raw is not None:
            # explicit empty string -> clear assignment
            if assigned_raw == '':
                updates.append("assigned_to = NULL")
                updates.append("assigned_type = NULL")
            else:
                assigned_to_val, assigned_type_val = None, None
                if ":" in assigned_raw:
                    # expected format "member:23" or "team:5"
                    assigned_type_val, assigned_to_val = assigned_raw.split(":", 1)
                else:
                    # fallback assume member id
                    assigned_type_val, assigned_to_val = "member", assigned_raw
                updates.append("assigned_to = %s")
                params.append(assigned_to_val)
                updates.append("assigned_type = %s")
                params.append(assigned_type_val)
            # remember for activity logging after update
            _assigned_change = assigned_raw
        else:
            _assigned_change = None

        # If frontend posts assigned_to_display, update a display column (you have used assigned_to_display in UI)
        assigned_to_display = request.POST.get('assigned_to_display')
        if assigned_to_display is not None:
            # Ensure your table has this column; if not, you can skip or store in a comment field
            # For compatibility with your DB schema above, store it in assigned_to (if you use member: or team:)
            # Here we just store it to a display column if it exists
            try:
                # attempt to update assigned_to_display column if present
                cur.execute("SHOW COLUMNS FROM tasks LIKE 'assigned_to_display'")
                col = cur.fetchone()
                if col:
                    updates.append("assigned_to_display = %s")
                    params.append(assigned_to_display)
                else:
                    # if column absent, optionally skip — or update assigned_to raw (not ideal)
                    pass
            except Exception:
                # ignore schema-check errors and skip
                pass

        status = request.POST.get('status')
        if status is not None:
            updates.append("status = %s")
            params.append(status)

        if not updates:
            return JsonResponse({'ok': False, 'error': 'no fields to update'}, status=400)

        # build and execute update
        set_clause = ", ".join(updates)
        params.append(tid)
        sql = f"UPDATE tasks SET {set_clause}, updated_at = NOW() WHERE id = %s"
        cur.execute(sql, tuple(params))

        # optional: log activity
        performed_by = request.session.get("user_id")
        try:
            cur.execute(
                "INSERT INTO activity_log (entity_type, entity_id, action, performed_by) VALUES (%s,%s,%s,%s)",
                ("task", tid, f"updated_via_api", performed_by),
            )
        except Exception:
            # don't fail entire operation if logging fails
            pass

        conn.commit()

        # If status changed to a closed/finished state and the closer is the assignee,
        # send a real-time pop notification to the assigner (created_by) only.
        try:
            closed_states = ('Closed', 'Finished', 'Completed')
            new_status = status
            # determine current user id (member or user session)
            cur_user = request.session.get('user_id') or request.session.get('member_id')

            assigned_type = existing.get('assigned_type') if isinstance(existing, dict) else None
            assigned_to = existing.get('assigned_to') if isinstance(existing, dict) else None
            creator_id = existing.get('created_by') if isinstance(existing, dict) else None
            task_title = existing.get('title') if isinstance(existing, dict) else None

            if new_status and new_status in closed_states and assigned_type == 'member' and assigned_to and cur_user and int(assigned_to) == int(cur_user):
                # Only notify if creator exists and is different from the closer
                if creator_id and int(creator_id) != int(cur_user):
                    tenant_key = request.session.get('tenant_id') or resolve_tenant_key_from_request(request)
                    # Build message
                    message = f"{task_title or 'Task'} has been marked as {new_status} by the assignee."
                    NotificationManager.send_notification(
                        tenant_key,
                        int(creator_id),
                        "Task Completed",
                        message,
                        notification_type='task',
                        link=f"/tasks/{tid}/view/",
                        created_by_id=cur_user,
                    )
        except Exception:
            # don't block API response if notification fails
            pass

        return JsonResponse({'ok': True})
    except Exception as e:
        # log or return reasonable error message
        conn.rollback()
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)
    finally:
        cur.close()

def task_board_view(request):
    status_columns = ["Open", "In Progress", "Review", "Blocked", "Closed"]
    conn = get_tenant_conn(request)
    cur = conn.cursor()
    cur.execute("SELECT id, name FROM projects ORDER BY name")
    projects = cur.fetchall()
    cur.execute("SELECT id, email, first_name, last_name FROM members ORDER BY first_name")
    members = cur.fetchall()
    cur.execute("SELECT id, name FROM teams ORDER BY name")
    teams = cur.fetchall()
    cur.close()
    return render(
        request,
        "core/task_board.html",
        {"page": "task_board", "status_columns": status_columns,
         "projects": projects, "members": members, "teams": teams},
    )


@require_GET
def api_get_subprojects(request):
    """
    API endpoint: GET /tasks/api/subprojects/?project_id=<id>
    Returns: { subprojects: [ {id, name}, ... ] }
    """
    project_id = request.GET.get("project_id")
    if not project_id:
        return JsonResponse({"error": "Missing project_id"}, status=400)
    conn = get_tenant_conn(request)
    cur = conn.cursor()
    try:
        cur.execute("SELECT id, name FROM subprojects WHERE project_id = %s ORDER BY name", (project_id,))
        rows = cur.fetchall()
        # handle both dict and tuple cursor
        subprojects = []
        if rows:
            if isinstance(rows[0], dict):
                subprojects = [{"id": r["id"], "name": r["name"]} for r in rows]
            else:
                subprojects = [{"id": r[0], "name": r[1]} for r in rows]
        return JsonResponse({"subprojects": subprojects})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)
    finally:
        cur.close()
    

def task_detail_view(request, task_id):
    conn = get_tenant_conn(request)
    cur = conn.cursor()
    cur.execute("SELECT id, title, description, status, priority, due_date, created_at FROM tasks WHERE id=%s", (task_id,))
    print("Executing SQL for task detail:", task_id)
    task = cur.fetchone()

    # Fetch timer history for this task
    cur.execute("""
        SELECT ts.id, ts.user_id, m.first_name, m.last_name, ts.start_time, ts.end_time, ts.duration_seconds, ts.notes
        FROM timer_sessions ts
        LEFT JOIN members m ON ts.user_id = m.id
        WHERE ts.task_id = %s
        ORDER BY ts.start_time DESC
    """, (task_id,))
    timer_history = cur.fetchall()
    cur.close()

    if not task:
        return render(request, "core/404.html", status=404)

    return render(request, "core/task_detail.html", {"task": task, "timer_history": timer_history})
def edit_task_view(request, task_id):
    conn = get_tenant_conn(request)
    cur = conn.cursor()

    if request.method == "POST":
        data = request.POST
        title = data.get("title")
        description = data.get("description")
        due_date = data.get("due_date") or None
        priority = data.get("priority") or "Normal"
        status = data.get("status") or "Open"

        #Automatic set closure date when status is 'Closed'
        if status == "Closed":
            closure_date=datetime.date.today()
        else:
            closure_date=None

        # Fetch existing meta before update so we can log and notify
        cur.execute("SELECT id, title, description, status, priority, due_date, created_by, assigned_to, assigned_type FROM tasks WHERE id=%s LIMIT 1", (task_id,))
        _existing = cur.fetchone()

        # Perform update
        cur.execute(
            """UPDATE tasks
               SET title=%s, description=%s, status=%s, priority=%s, due_date=%s, closure_date=%s, updated_at=NOW()
               WHERE id=%s""",
            (title, description, status, priority, due_date, closure_date, task_id),
        )

        # Log changes to activity_log for timeline
        try:
            performed_by = request.session.get('user_id') or request.session.get('member_id')
            if _existing:
                # Compare and log status change
                old_status = _existing.get('status') if isinstance(_existing, dict) else None
                if old_status != status:
                    cur.execute(
                        "INSERT INTO activity_log (entity_type, entity_id, action, performed_by, timestamp) VALUES (%s,%s,%s,%s,NOW())",
                        ("task", task_id, f"Changed status from {old_status} to {status}", performed_by),
                    )

                # Compare and log priority change
                old_priority = _existing.get('priority') if isinstance(_existing, dict) else None
                if old_priority != priority:
                    cur.execute(
                        "INSERT INTO activity_log (entity_type, entity_id, action, performed_by, timestamp) VALUES (%s,%s,%s,%s,NOW())",
                        ("task", task_id, f"Changed priority from {old_priority} to {priority}", performed_by),
                    )

                # Title change
                old_title = _existing.get('title') if isinstance(_existing, dict) else None
                if old_title != title:
                    cur.execute(
                        "INSERT INTO activity_log (entity_type, entity_id, action, performed_by, timestamp) VALUES (%s,%s,%s,%s,NOW())",
                        ("task", task_id, f"Changed title from {old_title} to {title}", performed_by),
                    )

                # Description change
                old_desc = _existing.get('description') if isinstance(_existing, dict) else None
                if (old_desc or '') != (description or ''):
                    cur.execute(
                        "INSERT INTO activity_log (entity_type, entity_id, action, performed_by, timestamp) VALUES (%s,%s,%s,%s,NOW())",
                        ("task", task_id, f"Updated description", performed_by),
                    )
        except Exception:
            pass

        conn.commit()

        # If status moved to closed and the closer is the assignee, notify the creator
        try:
            closed_states = ('Closed', 'Finished', 'Completed')
            cur_user = request.session.get('user_id') or request.session.get('member_id')
            if status in closed_states and _existing:
                assigned_type = _existing.get('assigned_type') if isinstance(_existing, dict) else None
                assigned_to = _existing.get('assigned_to') if isinstance(_existing, dict) else None
                creator_id = _existing.get('created_by') if isinstance(_existing, dict) else None
                task_title = _existing.get('title') if isinstance(_existing, dict) else None
                if assigned_type == 'member' and assigned_to and cur_user and int(assigned_to) == int(cur_user):
                    if creator_id and int(creator_id) != int(cur_user):
                        tenant_key = request.session.get('tenant_id') or resolve_tenant_key_from_request(request)
                        message = f"{task_title or 'Task'} has been marked as {status} by the assignee."
                        NotificationManager.send_notification(
                            tenant_key,
                            int(creator_id),
                            "Task Completed",
                            message,
                            notification_type='task',
                            link=f"/tasks/{task_id}/view/",
                            created_by_id=cur_user,
                        )
        except Exception:
            pass

        # Re-fetch updated task
        cur.execute("SELECT id, title, description, status, priority, due_date, created_at FROM tasks WHERE id=%s", (task_id,))
        task = cur.fetchone()
        cur.close()
        if not task:
            return render(request, "core/404.html", status=404)
        return redirect("my_tasks")

    # GET
    cur.execute("SELECT id, title, description, status, priority, due_date, created_at FROM tasks WHERE id=%s", (task_id,))
    row = cur.fetchone()
    if not row:
        cur.close()
        return render(request, "core/404.html", status=404)

    # Ensure row is a dict for template compatibility and description is not None
    if isinstance(row, dict):
        task = row
    else:
        cols = [desc[0] for desc in cur.description]
        task = dict(zip(cols, row))
    cur.close()

    # Fix: If description is None, set to empty string
    if task.get('description') is None:
        task['description'] = ''

    return render(request, "core/edit_task.html", {"task": task})

@require_POST
def delete_task_view(request, task_id): 
    """Deletes the specified task."""
    conn = get_tenant_conn(request)
    cur = conn.cursor()

    # Check if task exists
    cur.execute("SELECT id FROM tasks WHERE id=%s", (task_id,))
    task = cur.fetchone()
    if not task:
        cur.close()
        return render(request, "core/404.html", status=404)

    # Delete the task
    cur.execute("DELETE FROM tasks WHERE id=%s", (task_id,))
    conn.commit()
    cur.close()

    return redirect("my_tasks")


def export_task_pdf(request, task_id):
    """
    Export task details to PDF with professional formatting
    """
    conn = get_tenant_conn(request)
    cur = conn.cursor()
    
    # Fetch task details
    cur.execute("""
        SELECT 
            t.id, t.title, t.description, t.status, t.priority, 
            t.due_date, t.created_at, t.assigned_to, t.assigned_type,
            CASE
                WHEN t.assigned_type = 'member' THEN (SELECT CONCAT(m.first_name, ' ', m.last_name) FROM members m WHERE m.id = t.assigned_to)
                WHEN t.assigned_type = 'team' THEN (SELECT tm.name FROM teams tm WHERE tm.id = t.assigned_to)
                ELSE NULL
            END AS assigned_to_display,
            (SELECT p.name FROM projects p WHERE p.id = t.project_id) AS project_name
        FROM tasks t 
        WHERE t.id=%s
    """, (task_id,))
    
    task = cur.fetchone()
    
    if not task:
        cur.close()
        return HttpResponse("Task not found", status=404)
    
    # Convert to dict if tuple
    if not isinstance(task, dict):
        cols = [desc[0] for desc in cur.description]
        task = dict(zip(cols, task))
    
    # Fetch timer history
    cur.execute("""
        SELECT ts.user_id, m.first_name, m.last_name, ts.start_time, 
               ts.end_time, ts.duration_seconds, ts.notes
        FROM timer_sessions ts
        LEFT JOIN members m ON ts.user_id = m.id
        WHERE ts.task_id = %s
        ORDER BY ts.start_time DESC
    """, (task_id,))
    timer_history = cur.fetchall()
    cur.close()
    
    # Create PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="task_{task_id}_{task.get("title", "details").replace(" ", "_")}.pdf"'
    
    # Create PDF document
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    
    # Container for PDF elements
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#3b82f6'),
        spaceAfter=12,
        spaceBefore=20,
        fontName='Helvetica-Bold'
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.HexColor('#1e293b'),
        spaceAfter=12
    )
    
    label_style = ParagraphStyle(
        'Label',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#64748b'),
        fontName='Helvetica-Bold',
        spaceAfter=4
    )
    
    # Add title
    elements.append(Paragraph(f"Task: {task.get('title', 'N/A')}", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Add status badge (simulated with colored text)
    status_text = f'<para align="center" backColor="#3b82f6" textColor="white" fontSize="10" fontName="Helvetica-Bold">&nbsp;&nbsp;{task.get("status", "N/A").upper()}&nbsp;&nbsp;</para>'
    elements.append(Paragraph(status_text, normal_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Create details table
    details_data = [
        ['DESCRIPTION', task.get('description', 'No description provided.')],
        ['DUE DATE', str(task.get('due_date', 'Not set'))],
        ['CREATED AT', str(task.get('created_at', 'N/A'))],
        ['PRIORITY', task.get('priority', 'Normal')],
        ['ASSIGNED TO', task.get('assigned_to_display', 'Unassigned')],
        ['PROJECT', task.get('project_name', 'N/A')]
    ]
    
    details_table = Table(details_data, colWidths=[2*inch, 4.5*inch])
    details_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f1f5f9')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#64748b')),
        ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#1e293b')),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    
    elements.append(details_table)
    elements.append(Spacer(1, 0.4*inch))
    
    # Add timer history section
    if timer_history and len(timer_history) > 0:
        elements.append(Paragraph("⏱️ TIMER HISTORY", heading_style))
        elements.append(Spacer(1, 0.1*inch))
        
        # Create timer table
        timer_data = [['User', 'Start Time', 'End Time', 'Duration', 'Notes']]
        for session in timer_history:
            if isinstance(session, dict):
                user = f"{session.get('first_name', '')} {session.get('last_name', '')}".strip() or str(session.get('user_id', 'N/A'))
                timer_data.append([
                    user,
                    str(session.get('start_time', 'N/A')),
                    str(session.get('end_time', '-')),
                    f"{session.get('duration_seconds', 0)}s",
                    str(session.get('notes', '-'))
                ])
            else:
                user = f"{session[1]} {session[2]}".strip() or str(session[0])
                timer_data.append([
                    user,
                    str(session[3]),
                    str(session[4] if session[4] else '-'),
                    f"{session[5]}s",
                    str(session[6] if session[6] else '-')
                ])
        
        timer_table = Table(timer_data, colWidths=[1.2*inch, 1.5*inch, 1.5*inch, 0.8*inch, 1.5*inch])
        timer_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f9fafb')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')])
        ]))
        
        elements.append(timer_table)
    else:
        elements.append(Paragraph("⏱️ TIMER HISTORY", heading_style))
        elements.append(Paragraph("No timer history found for this task.", normal_style))
    
    # Build PDF
    doc.build(elements)
    
    # Get PDF from buffer
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response


# ==============================
#  API: GET PROJECT WORK TYPES
# ==============================
@require_GET
def api_get_project_work_types(request):
    """
    API endpoint to get configured work types for a project
    """
    project_id = request.GET.get('project_id')
    if not project_id:
        return JsonResponse({'work_types': []})
    
    conn = get_tenant_conn(request)
    cur = conn.cursor()
    
    try:
        # Get configured work types for the project
        cur.execute("""
            SELECT work_type FROM project_work_types 
            WHERE project_id = %s AND is_enabled = 1
        """, (project_id,))
        rows = cur.fetchall()
        
        if rows:
            # Map database values to display names
            type_mapping = {
                'task': 'Task',
                'bug': 'Bug',
                'story': 'Story',
                'defect': 'Defect',
                'subtask': 'Sub Task',
                'report': 'Report',
                'change_request': 'Change Request'
            }
            work_types = [type_mapping.get(row['work_type'], row['work_type']) for row in rows]
        else:
            # If no configuration exists, return all types
            work_types = ['Task', 'Bug', 'Story', 'Defect', 'Sub Task', 'Report', 'Change Request']
        
        return JsonResponse({'work_types': work_types})
    except Exception as e:
        return JsonResponse({'work_types': [], 'error': str(e)})
    finally:
        cur.close()


# ==============================
#  CREATE BUG
# ==============================
def create_bug_view(request):
    """
    Creates a new bug with bug-specific fields
    """
    # Check if Bug work type is enabled for this tenant
    enabled_work_types = get_tenant_work_types(request)
    if 'Bug' not in enabled_work_types:
        from django.contrib import messages
        messages.error(request, "Bug work type is not enabled for your organization.")
        return redirect("task_board")
    
    conn = get_tenant_conn(request)
    cur = conn.cursor()

    if request.method == "POST":
        data = request.POST
        project_id = data.get("project_id") or None
        subproject_id = data.get("subproject_id") or None
        title = data.get("title")
        description = data.get("description")
        steps_to_reproduce = data.get("steps_to_reproduce")
        expected_behavior = data.get("expected_behavior")
        actual_behavior = data.get("actual_behavior")
        severity = data.get("severity") or "Medium"
        due_date = data.get("due_date") or None
        closure_date = data.get("closure_date") or None
        priority = data.get("priority") or "Normal"
        status = data.get("status") or "Open"
        work_type = "Bug"
        created_by = request.session.get("user_id")
        
        # Capture system information
        si_browser = data.get("si_browser") or None
        si_resolution = data.get("si_resolution") or None
        si_os = data.get("si_os") or None
        si_timestamp = data.get("si_timestamp") or None

        # Combine bug-specific fields into description
        full_description = f"{description}\n\n"
        if steps_to_reproduce:
            full_description += f"**Steps to Reproduce:**\n{steps_to_reproduce}\n\n"
        if expected_behavior:
            full_description += f"**Expected Behavior:**\n{expected_behavior}\n\n"
        if actual_behavior:
            full_description += f"**Actual Behavior:**\n{actual_behavior}\n\n"
        
        # Handle assignment
        assigned_raw = data.get("assigned_to") or None
        assigned_to, assigned_type = None, None
        if assigned_raw:
            if ":" in assigned_raw:
                assigned_type, assigned_to = assigned_raw.split(":", 1)
            else:
                assigned_type, assigned_to = "member", assigned_raw

        # INSERT
        cur.execute(
            """INSERT INTO tasks
               (project_id, subproject_id, title, description, status, priority,
                assigned_to, assigned_type, created_by, due_date, closure_date, work_type,
                si_browser, si_resolution, si_os, si_timestamp, created_at)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())""",
            (
                project_id,
                subproject_id,
                title,
                full_description,
                status,
                severity,  # Use severity as priority for bugs
                assigned_to,
                assigned_type,
                created_by,
                due_date,
                closure_date,
                work_type,
                si_browser,
                si_resolution,
                si_os,
                si_timestamp,
            ),
        )
        task_id = cur.lastrowid

        # Log activity
        cur.execute(
            "INSERT INTO activity_log (entity_type, entity_id, action, performed_by) VALUES (%s,%s,%s,%s)",
            ("task", task_id, "created", created_by),
        )

        # Create notification if task is assigned to a member
        if assigned_type == "member" and assigned_to:
            cur.execute("SELECT CONCAT(first_name, ' ', last_name) as name FROM members WHERE id=%s", (created_by,))
            creator = cur.fetchone()
            creator_name = creator['name'] if creator else 'Someone'
            
            cur.execute("""
                INSERT INTO notifications (user_id, title, message, type, link)
                VALUES (%s, %s, %s, %s, %s)
            """, (
                assigned_to,
                "New Bug Assigned",
                f"{creator_name} assigned you a bug: '{title}'",
                "task",
                f"/tasks/{task_id}/view/"
            ))

        # Handle file attachments using helper function
        save_task_attachments(request, task_id, cur, created_by)

        conn.commit()
        cur.close()
        return redirect("task_board")

    # GET
    cur.execute("SELECT id, name FROM projects ORDER BY name")
    projects = cur.fetchall()
    cur.execute("SELECT id, email, first_name, last_name FROM members ORDER BY first_name")
    members = cur.fetchall()
    cur.execute("SELECT id, name FROM teams ORDER BY name")
    teams = cur.fetchall()
    
    cur.close()

    return render(
        request,
        "core/create_bug.html",
        {
            "projects": projects, 
            "members": members, 
            "teams": teams, 
            "page": "task_create"
        },
    )


# ==============================
#  CREATE STORY
# ==============================
def create_story_view(request):
    """
    Creates a new user story with story-specific fields
    """
    # Check if Story work type is enabled for this tenant
    enabled_work_types = get_tenant_work_types(request)
    if 'Story' not in enabled_work_types:
        from django.contrib import messages
        messages.error(request, "Story work type is not enabled for your organization.")
        return redirect("task_board")
    
    conn = get_tenant_conn(request)
    cur = conn.cursor()

    if request.method == "POST":
        data = request.POST
        project_id = data.get("project_id") or None
        subproject_id = data.get("subproject_id") or None
        title = data.get("title")
        user_story = data.get("user_story")
        acceptance_criteria = data.get("acceptance_criteria")
        story_points = data.get("story_points") or None
        due_date = data.get("due_date") or None
        closure_date = data.get("closure_date") or None
        priority = data.get("priority") or "Normal"
        status = data.get("status") or "Open"
        work_type = "Story"
        created_by = request.session.get("user_id")
        
        # Capture system information
        si_browser = data.get("si_browser") or None
        si_resolution = data.get("si_resolution") or None
        si_os = data.get("si_os") or None
        si_timestamp = data.get("si_timestamp") or None

        # Combine story-specific fields into description
        full_description = ""
        if user_story:
            full_description += f"**User Story:**\n{user_story}\n\n"
        if acceptance_criteria:
            full_description += f"**Acceptance Criteria:**\n{acceptance_criteria}\n\n"
        if story_points:
            full_description += f"**Story Points:** {story_points}\n\n"
        
        # Handle assignment
        assigned_raw = data.get("assigned_to") or None
        assigned_to, assigned_type = None, None
        if assigned_raw:
            if ":" in assigned_raw:
                assigned_type, assigned_to = assigned_raw.split(":", 1)
            else:
                assigned_type, assigned_to = "member", assigned_raw

        # INSERT
        cur.execute(
            """INSERT INTO tasks
               (project_id, subproject_id, title, description, status, priority,
                assigned_to, assigned_type, created_by, due_date, closure_date, work_type,
                si_browser, si_resolution, si_os, si_timestamp, created_at)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())""",
            (
                project_id,
                subproject_id,
                title,
                full_description,
                status,
                priority,
                assigned_to,
                assigned_type,
                created_by,
                due_date,
                closure_date,
                work_type,
                si_browser,
                si_resolution,
                si_os,
                si_timestamp,
            ),
        )
        task_id = cur.lastrowid

        # Log activity
        cur.execute(
            "INSERT INTO activity_log (entity_type, entity_id, action, performed_by) VALUES (%s,%s,%s,%s)",
            ("task", task_id, "created", created_by),
        )

        # Create notification if task is assigned to a member
        if assigned_type == "member" and assigned_to:
            cur.execute("SELECT CONCAT(first_name, ' ', last_name) as name FROM members WHERE id=%s", (created_by,))
            creator = cur.fetchone()
            creator_name = creator['name'] if creator else 'Someone'
            
            cur.execute("""
                INSERT INTO notifications (user_id, title, message, type, link)
                VALUES (%s, %s, %s, %s, %s)
            """, (
                assigned_to,
                "New Story Assigned",
                f"{creator_name} assigned you a story: '{title}'",
                "task",
                f"/tasks/{task_id}/view/"
            ))

        # Handle file attachments using helper function
        save_task_attachments(request, task_id, cur, created_by)

        conn.commit()
        cur.close()
        return redirect("task_board")

    # GET
    cur.execute("SELECT id, name FROM projects ORDER BY name")
    projects = cur.fetchall()
    cur.execute("SELECT id, email, first_name, last_name FROM members ORDER BY first_name")
    members = cur.fetchall()
    cur.execute("SELECT id, name FROM teams ORDER BY name")
    teams = cur.fetchall()
    
    cur.close()

    return render(
        request,
        "core/create_story.html",
        {
            "projects": projects, 
            "members": members, 
            "teams": teams, 
            "page": "task_create"
        },
    )


# ==============================
#  CREATE DEFECT
# ==============================
def create_defect_view(request):
    """
    Creates a new defect with defect-specific fields
    """
    # Check if Defect work type is enabled for this tenant
    enabled_work_types = get_tenant_work_types(request)
    if 'Defect' not in enabled_work_types:
        from django.contrib import messages
        messages.error(request, "Defect work type is not enabled for your organization.")
        return redirect("task_board")
    
    conn = get_tenant_conn(request)
    cur = conn.cursor()

    if request.method == "POST":
        data = request.POST
        project_id = data.get("project_id") or None
        subproject_id = data.get("subproject_id") or None
        title = data.get("title")
        description = data.get("description")
        environment = data.get("environment")
        impact = data.get("impact") or "Medium"
        due_date = data.get("due_date") or None
        closure_date = data.get("closure_date") or None
        priority = data.get("priority") or "Normal"
        status = data.get("status") or "Open"
        work_type = "Defect"
        created_by = request.session.get("user_id")
        
        # Capture system information
        si_browser = data.get("si_browser") or None
        si_resolution = data.get("si_resolution") or None
        si_os = data.get("si_os") or None
        si_timestamp = data.get("si_timestamp") or None

        # Combine defect-specific fields into description
        full_description = f"{description}\n\n"
        if environment:
            full_description += f"**Environment:**\n{environment}\n\n"
        full_description += f"**Impact:** {impact}\n\n"
        
        # Handle assignment
        assigned_raw = data.get("assigned_to") or None
        assigned_to, assigned_type = None, None
        if assigned_raw:
            if ":" in assigned_raw:
                assigned_type, assigned_to = assigned_raw.split(":", 1)
            else:
                assigned_type, assigned_to = "member", assigned_raw

        # INSERT
        cur.execute(
            """INSERT INTO tasks
               (project_id, subproject_id, title, description, status, priority,
                assigned_to, assigned_type, created_by, due_date, closure_date, work_type,
                si_browser, si_resolution, si_os, si_timestamp, created_at)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())""",
            (
                project_id,
                subproject_id,
                title,
                full_description,
                status,
                priority,
                assigned_to,
                assigned_type,
                created_by,
                due_date,
                closure_date,
                work_type,
                si_browser,
                si_resolution,
                si_os,
                si_timestamp,
            ),
        )
        task_id = cur.lastrowid

        # Log activity
        cur.execute(
            "INSERT INTO activity_log (entity_type, entity_id, action, performed_by) VALUES (%s,%s,%s,%s)",
            ("task", task_id, "created", created_by),
        )

        # Create notification if task is assigned to a member
        if assigned_type == "member" and assigned_to:
            cur.execute("SELECT CONCAT(first_name, ' ', last_name) as name FROM members WHERE id=%s", (created_by,))
            creator = cur.fetchone()
            creator_name = creator['name'] if creator else 'Someone'
            
            cur.execute("""
                INSERT INTO notifications (user_id, title, message, type, link)
                VALUES (%s, %s, %s, %s, %s)
            """, (
                assigned_to,
                "New Defect Assigned",
                f"{creator_name} assigned you a defect: '{title}'",
                "task",
                f"/tasks/{task_id}/view/"
            ))

        # Handle file attachments using helper function
        save_task_attachments(request, task_id, cur, created_by)

        conn.commit()
        cur.close()
        return redirect("task_board")

    # GET
    cur.execute("SELECT id, name FROM projects ORDER BY name")
    projects = cur.fetchall()
    cur.execute("SELECT id, email, first_name, last_name FROM members ORDER BY first_name")
    members = cur.fetchall()
    cur.execute("SELECT id, name FROM teams ORDER BY name")
    teams = cur.fetchall()
    
    cur.close()

    return render(
        request,
        "core/create_defect.html",
        {
            "projects": projects, 
            "members": members, 
            "teams": teams, 
            "page": "task_create"
        },
    )


# ==============================
#  CREATE SUB TASK
# ==============================
def create_subtask_view(request):
    """
    Creates a new sub task with parent task selection
    """
    # Check if Sub Task work type is enabled for this tenant
    enabled_work_types = get_tenant_work_types(request)
    if 'Sub Task' not in enabled_work_types:
        from django.contrib import messages
        messages.error(request, "Sub Task work type is not enabled for your organization.")
        return redirect("task_board")
    
    conn = get_tenant_conn(request)
    cur = conn.cursor()

    if request.method == "POST":
        data = request.POST
        parent_task_id = data.get("parent_task_id") or None
        project_id = data.get("project_id") or None
        subproject_id = data.get("subproject_id") or None
        title = data.get("title")
        description = data.get("description")
        estimated_hours = data.get("estimated_hours") or None
        due_date = data.get("due_date") or None
        closure_date = data.get("closure_date") or None
        priority = data.get("priority") or "Normal"
        status = data.get("status") or "Open"
        work_type = "Sub Task"
        created_by = request.session.get("user_id")
        
        # Capture system information
        si_browser = data.get("si_browser") or None
        si_resolution = data.get("si_resolution") or None
        si_os = data.get("si_os") or None
        si_timestamp = data.get("si_timestamp") or None

        # Add parent task info to description
        full_description = f"{description}\n\n"
        if parent_task_id:
            cur.execute("SELECT title FROM tasks WHERE id=%s", (parent_task_id,))
            parent = cur.fetchone()
            if parent:
                full_description += f"**Parent Task:** {parent['title']} (ID: {parent_task_id})\n\n"
        if estimated_hours:
            full_description += f"**Estimated Hours:** {estimated_hours}\n\n"
        
        # Handle assignment
        assigned_raw = data.get("assigned_to") or None
        assigned_to, assigned_type = None, None
        if assigned_raw:
            if ":" in assigned_raw:
                assigned_type, assigned_to = assigned_raw.split(":", 1)
            else:
                assigned_type, assigned_to = "member", assigned_raw

        # INSERT
        cur.execute(
            """INSERT INTO tasks
               (project_id, subproject_id, title, description, status, priority,
                assigned_to, assigned_type, created_by, due_date, closure_date, work_type,
                si_browser, si_resolution, si_os, si_timestamp, created_at)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())""",
            (
                project_id,
                subproject_id,
                title,
                full_description,
                status,
                priority,
                assigned_to,
                assigned_type,
                created_by,
                due_date,
                closure_date,
                work_type,
                si_browser,
                si_resolution,
                si_os,
                si_timestamp,
            ),
        )
        task_id = cur.lastrowid

        # Log activity
        cur.execute(
            "INSERT INTO activity_log (entity_type, entity_id, action, performed_by) VALUES (%s,%s,%s,%s)",
            ("task", task_id, "created", created_by),
        )

        # Create notification if task is assigned to a member
        if assigned_type == "member" and assigned_to:
            cur.execute("SELECT CONCAT(first_name, ' ', last_name) as name FROM members WHERE id=%s", (created_by,))
            creator = cur.fetchone()
            creator_name = creator['name'] if creator else 'Someone'
            
            cur.execute("""
                INSERT INTO notifications (user_id, title, message, type, link)
                VALUES (%s, %s, %s, %s, %s)
            """, (
                assigned_to,
                "New Sub Task Assigned",
                f"{creator_name} assigned you a sub task: '{title}'",
                "task",
                f"/tasks/{task_id}/view/"
            ))

        # Handle file attachments using helper function
        save_task_attachments(request, task_id, cur, created_by)

        conn.commit()
        cur.close()
        return redirect("task_board")

    # GET
    cur.execute("SELECT id, name FROM projects ORDER BY name")
    projects = cur.fetchall()
    cur.execute("SELECT id, email, first_name, last_name FROM members ORDER BY first_name")
    members = cur.fetchall()
    cur.execute("SELECT id, name FROM teams ORDER BY name")
    teams = cur.fetchall()
    cur.execute("SELECT id, title FROM tasks WHERE work_type != 'Sub Task' ORDER BY created_at DESC")
    parent_tasks = cur.fetchall()
    
    cur.close()

    return render(
        request,
        "core/create_subtask.html",
        {
            "projects": projects, 
            "members": members, 
            "teams": teams, 
            "parent_tasks": parent_tasks,
            "page": "task_create"
        },
    )


# ==============================
#  CREATE REPORT
# ==============================
def create_report_view(request):
    """
    Creates a new report task
    """
    # Check if Report work type is enabled for this tenant
    enabled_work_types = get_tenant_work_types(request)
    if 'Report' not in enabled_work_types:
        from django.contrib import messages
        messages.error(request, "Report work type is not enabled for your organization.")
        return redirect("task_board")
    
    conn = get_tenant_conn(request)
    cur = conn.cursor()

    if request.method == "POST":
        data = request.POST
        project_id = data.get("project_id") or None
        subproject_id = data.get("subproject_id") or None
        title = data.get("title")
        report_type = data.get("report_type") or "Status Report"
        frequency = data.get("frequency") or "One-time"
        data_sources = data.get("data_sources")
        recipients = data.get("recipients")
        description = data.get("description")
        due_date = data.get("due_date") or None
        closure_date = data.get("closure_date") or None
        priority = data.get("priority") or "Normal"
        status = data.get("status") or "Open"
        work_type = "Report"
        created_by = request.session.get("user_id")
        
        # Capture system information
        si_browser = data.get("si_browser") or None
        si_resolution = data.get("si_resolution") or None
        si_os = data.get("si_os") or None
        si_timestamp = data.get("si_timestamp") or None

        # Build report description
        full_description = f"**Report Type:** {report_type}\n\n"
        full_description += f"**Frequency:** {frequency}\n\n"
        if data_sources:
            full_description += f"**Data Sources:**\n{data_sources}\n\n"
        if recipients:
            full_description += f"**Recipients:**\n{recipients}\n\n"
        if description:
            full_description += f"**Additional Details:**\n{description}\n\n"
        
        # Handle assignment
        assigned_raw = data.get("assigned_to") or None
        assigned_to, assigned_type = None, None
        if assigned_raw:
            if ":" in assigned_raw:
                assigned_type, assigned_to = assigned_raw.split(":", 1)
            else:
                assigned_type, assigned_to = "member", assigned_raw

        # INSERT
        cur.execute(
            """INSERT INTO tasks
               (project_id, subproject_id, title, description, status, priority,
                assigned_to, assigned_type, created_by, due_date, closure_date, work_type,
                si_browser, si_resolution, si_os, si_timestamp, created_at)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())""",
            (
                project_id,
                subproject_id,
                title,
                full_description,
                status,
                priority,
                assigned_to,
                assigned_type,
                created_by,
                due_date,
                closure_date,
                work_type,
                si_browser,
                si_resolution,
                si_os,
                si_timestamp,
            ),
        )
        task_id = cur.lastrowid

        # Log activity
        cur.execute(
            "INSERT INTO activity_log (entity_type, entity_id, action, performed_by) VALUES (%s,%s,%s,%s)",
            ("task", task_id, "created", created_by),
        )

        # Create notification if task is assigned to a member
        if assigned_type == "member" and assigned_to:
            cur.execute("SELECT CONCAT(first_name, ' ', last_name) as name FROM members WHERE id=%s", (created_by,))
            creator = cur.fetchone()
            creator_name = creator['name'] if creator else 'Someone'
            
            cur.execute("""
                INSERT INTO notifications (user_id, title, message, type, link)
                VALUES (%s, %s, %s, %s, %s)
            """, (
                assigned_to,
                "New Report Task Assigned",
                f"{creator_name} assigned you a report: '{title}'",
                "task",
                f"/tasks/detail/{task_id}"
            ))

        conn.commit()
        cur.close()
        return redirect("task_board")

    # GET
    cur.execute("SELECT id, name FROM projects ORDER BY name")
    projects = cur.fetchall()
    cur.execute("SELECT id, email, first_name, last_name FROM members ORDER BY first_name")
    members = cur.fetchall()
    cur.execute("SELECT id, name FROM teams ORDER BY name")
    teams = cur.fetchall()
    
    cur.close()

    return render(
        request,
        "core/create_report.html",
        {
            "projects": projects, 
            "members": members, 
            "teams": teams, 
            "page": "task_create"
        },
    )


# ==============================
#  CREATE CHANGE REQUEST
# ==============================
def create_change_request_view(request):
    """
    Creates a new change request
    """
    # Check if Change Request work type is enabled for this tenant
    enabled_work_types = get_tenant_work_types(request)
    if 'Change Request' not in enabled_work_types:
        from django.contrib import messages
        messages.error(request, "Change Request work type is not enabled for your organization.")
        return redirect("task_board")
    
    conn = get_tenant_conn(request)
    cur = conn.cursor()

    if request.method == "POST":
        data = request.POST
        project_id = data.get("project_id") or None
        subproject_id = data.get("subproject_id") or None
        title = data.get("title")
        change_type = data.get("change_type") or "Feature"
        reason = data.get("reason")
        current_state = data.get("current_state")
        proposed_change = data.get("proposed_change")
        impact_analysis = data.get("impact_analysis")
        rollback_plan = data.get("rollback_plan")
        due_date = data.get("due_date") or None
        closure_date = data.get("closure_date") or None
        priority = data.get("priority") or "Normal"
        status = data.get("status") or "Pending Approval"
        work_type = "Change Request"
        created_by = request.session.get("user_id")
        
        # Capture system information
        si_browser = data.get("si_browser") or None
        si_resolution = data.get("si_resolution") or None
        si_os = data.get("si_os") or None
        si_timestamp = data.get("si_timestamp") or None

        # Build change request description
        full_description = f"**Change Type:** {change_type}\n\n"
        if reason:
            full_description += f"**Reason for Change:**\n{reason}\n\n"
        if current_state:
            full_description += f"**Current State:**\n{current_state}\n\n"
        if proposed_change:
            full_description += f"**Proposed Change:**\n{proposed_change}\n\n"
        if impact_analysis:
            full_description += f"**Impact Analysis:**\n{impact_analysis}\n\n"
        if rollback_plan:
            full_description += f"**Rollback Plan:**\n{rollback_plan}\n\n"
        
        # Handle assignment
        assigned_raw = data.get("assigned_to") or None
        assigned_to, assigned_type = None, None
        if assigned_raw:
            if ":" in assigned_raw:
                assigned_type, assigned_to = assigned_raw.split(":", 1)
            else:
                assigned_type, assigned_to = "member", assigned_raw

        # INSERT
        cur.execute(
            """INSERT INTO tasks
               (project_id, subproject_id, title, description, status, priority,
                assigned_to, assigned_type, created_by, due_date, closure_date, work_type,
                si_browser, si_resolution, si_os, si_timestamp, created_at)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())""",
            (
                project_id,
                subproject_id,
                title,
                full_description,
                status,
                priority,
                assigned_to,
                assigned_type,
                created_by,
                due_date,
                closure_date,
                work_type,
                si_browser,
                si_resolution,
                si_os,
                si_timestamp,
            ),
        )
        task_id = cur.lastrowid

        # Log activity
        cur.execute(
            "INSERT INTO activity_log (entity_type, entity_id, action, performed_by) VALUES (%s,%s,%s,%s)",
            ("task", task_id, "created", created_by),
        )

        # Create notification if task is assigned to a member
        if assigned_type == "member" and assigned_to:
            cur.execute("SELECT CONCAT(first_name, ' ', last_name) as name FROM members WHERE id=%s", (created_by,))
            creator = cur.fetchone()
            creator_name = creator['name'] if creator else 'Someone'
            
            cur.execute("""
                INSERT INTO notifications (user_id, title, message, type, link)
                VALUES (%s, %s, %s, %s, %s)
            """, (
                assigned_to,
                "New Change Request Assigned",
                f"{creator_name} assigned you a change request: '{title}'",
                "task",
                f"/tasks/detail/{task_id}"
            ))

        # Handle file attachments using helper function
        save_task_attachments(request, task_id, cur, created_by)

        conn.commit()
        cur.close()
        return redirect("task_board")

    # GET
    cur.execute("SELECT id, name FROM projects ORDER BY name")
    projects = cur.fetchall()
    cur.execute("SELECT id, email, first_name, last_name FROM members ORDER BY first_name")
    members = cur.fetchall()
    cur.execute("SELECT id, name FROM teams ORDER BY name")
    teams = cur.fetchall()
    
    cur.close()

    return render(
        request,
        "core/create_change_request.html",
        {
            "projects": projects, 
            "members": members, 
            "teams": teams, 
            "page": "task_create"
        },
    )


# ==============================
#  TASK PAGE VIEW (Jira-style detail page)
# ==============================
def task_page_view(request, task_id):
    """
    Display a detailed Jira-style view of a task with:
    - Description
    - Comments
    - Activity Timeline
    - Details (Reporter, Assignee, Module, Resolution, Dates)
    - Environment (Browser, OS, Resolution)
    - Actions (Change Status, Update Priority, Add Attachment, Delete)
    """
    conn = get_tenant_conn(request)
    cur = conn.cursor()
    
    # Check if system info columns exist
    cur.execute("SHOW COLUMNS FROM tasks LIKE 'si_browser'")
    has_system_info = cur.fetchone() is not None
    
    # Build query based on available columns
    system_info_fields = ""
    if has_system_info:
        system_info_fields = "t.si_browser, t.si_resolution, t.si_os, t.si_timestamp,"
    
    # Get task details with all system information
    query = f"""
        SELECT 
            t.id, t.title, t.description, t.priority, t.status, t.work_type,
            t.due_date, t.closure_date, t.created_at, t.created_by,
            t.project_id, t.subproject_id,
            t.assigned_type, t.assigned_to,
            {system_info_fields}
            p.name AS project_name,
            sp.name AS subproject_name,
            CONCAT(creator.first_name, ' ', creator.last_name) AS creator_name,
            creator.email AS creator_email,
            CASE 
                WHEN t.assigned_type = 'member' THEN CONCAT(assignee.first_name, ' ', assignee.last_name)
                WHEN t.assigned_type = 'team' THEN tm.name
                ELSE NULL
            END AS assignee_name,
            CASE 
                WHEN t.assigned_type = 'member' THEN assignee.email
                ELSE NULL
            END AS assignee_email,
            tm.name AS team_name
        FROM tasks t
        LEFT JOIN projects p ON t.project_id = p.id
        LEFT JOIN subprojects sp ON t.subproject_id = sp.id
        LEFT JOIN members creator ON t.created_by = creator.id
        LEFT JOIN members assignee ON t.assigned_type = 'member' AND t.assigned_to = assignee.id
        LEFT JOIN teams tm ON t.assigned_type = 'team' AND t.assigned_to = tm.id
        WHERE t.id = %s
    """
    
    cur.execute(query, (task_id,))
    
    issue = cur.fetchone()
    
    if not issue:
        cur.close()
        return render(request, 'core/404.html', {'message': 'Task not found'}, status=404)
    
    # Prepare issue data for template
    issue_data = {
        'id': issue['id'],
        'issue_id': f"TASK-{issue['id']:04d}",  # Format as TASK-0001
        'summary': issue['title'],
        'description': issue['description'],
        'priority': issue['priority'],
        'status': issue['status'],
        'issue_type': issue['work_type'] or 'Task',
        'severity': 'BLOCKER' if issue['priority'] == 'Critical' else 'MAJOR' if issue['priority'] == 'High' else 'MINOR',
        'module': issue['project_name'],
        'resolution': 'PENDING' if issue['status'] in ['Open', 'New'] else 'FINISHED' if issue['status'] == 'Closed' else 'IN_PROGRESS',
        'created_at': issue['created_at'],
        'due_date': issue['due_date'],
        'closed_date': issue['closure_date'],
        'start_date': None,  # Can be added if you have a start_date field
        'reporter_name': issue['creator_name'],
        'reporter_email': issue['creator_email'],
        'reporter_ldap': issue['creator_email'],
        'assignee_name': issue['assignee_name'] or issue['team_name'],
        'assignee_ldap': issue['assignee_email'] or f"Team: {issue['team_name']}" if issue['team_name'] else None,
        'browser_info': issue.get('si_browser') if has_system_info else None,
        'screen_resolution': issue.get('si_resolution') if has_system_info else None,
        'os_info': issue.get('si_os') if has_system_info else None,
    }

    # If assignee email is missing (e.g. stored inconsistently), try to resolve it from members table
    if not issue_data.get('assignee_ldap') and issue_data.get('assignee_name'):
        try:
            # try exact full-name match
            cur.execute(
                "SELECT email FROM members WHERE CONCAT(first_name, ' ', last_name) = %s LIMIT 1",
                (issue_data['assignee_name'],),
            )
            mem = cur.fetchone()
            if mem and mem.get('email'):
                issue_data['assignee_ldap'] = mem['email']
            else:
                # try matching by first name (best-effort)
                parts = issue_data['assignee_name'].split()
                if parts:
                    cur.execute("SELECT email FROM members WHERE first_name = %s LIMIT 1", (parts[0],))
                    mem = cur.fetchone()
                    if mem and mem.get('email'):
                        issue_data['assignee_ldap'] = mem['email']
        except Exception:
            # don't fail rendering if lookup errors
            pass

    # Normalize missing values to empty string so template doesn't print 'None'
    if issue_data.get('assignee_ldap') in (None, 'None'):
        issue_data['assignee_ldap'] = ''
    
    # Get comments from task_comments table
    cur.execute("""
        SELECT 
            id, comment_text, commenter_id, commenter_name, 
            is_internal, created_at
        FROM task_comments
        WHERE task_id = %s
        ORDER BY created_at ASC
    """, (task_id,))
    
    comments = cur.fetchall()
    
    # Get activity timeline from activity_log table
    activities = []
    
    # Add task creation activity
    activities.append({
        'user_name': issue['creator_name'],
        'action_type': 'created',
        'old_value': None,
        'new_value': None,
        'description': f"Issue {issue_data['issue_id']} created",
        'created_at': issue['created_at']
    })
    
    # If assigned, add assignment activity
    if issue['assignee_name'] or issue['team_name']:
        assignee_text = issue['assignee_name'] or f"Team: {issue['team_name']}"
        activities.append({
            'user_name': issue['creator_name'],
            'action_type': 'assigned',
            'old_value': 'Unassigned',
            'new_value': assignee_text,
            'description': f"Assigned to {assignee_text}",
            'created_at': issue['created_at']
        })
    
    # Get activities from activity_log table
    cur.execute("""
        SELECT 
            al.action,
            al.timestamp as created_at,
            CONCAT(m.first_name, ' ', m.last_name) as user_name
        FROM activity_log al
        LEFT JOIN members m ON al.performed_by = m.id
        WHERE al.entity_type = 'task' AND al.entity_id = %s
        ORDER BY al.timestamp ASC
    """, (task_id,))
    
    log_activities = cur.fetchall()
    
    # Add logged activities to timeline
    for log_activity in log_activities:
        action_text = log_activity['action']
        
        # Parse status/priority changes
        if 'status from' in action_text.lower():
            parts = action_text.split(' from ')
            if len(parts) == 2:
                old_new = parts[1].split(' to ')
                activities.append({
                    'user_name': log_activity['user_name'],
                    'action_type': 'status_changed',
                    'old_value': old_new[0].strip() if len(old_new) > 0 else None,
                    'new_value': old_new[1].strip() if len(old_new) > 1 else None,
                    'description': action_text,
                    'created_at': log_activity['created_at']
                })
        elif 'priority from' in action_text.lower():
            parts = action_text.split(' from ')
            if len(parts) == 2:
                old_new = parts[1].split(' to ')
                activities.append({
                    'user_name': log_activity['user_name'],
                    'action_type': 'priority_changed',
                    'old_value': old_new[0].strip() if len(old_new) > 0 else None,
                    'new_value': old_new[1].strip() if len(old_new) > 1 else None,
                    'description': action_text,
                    'created_at': log_activity['created_at']
                })
        elif 'added comment' in action_text.lower() or 'added internal comment' in action_text.lower():
            activities.append({
                'user_name': log_activity['user_name'],
                'action_type': 'commented',
                'old_value': None,
                'new_value': None,
                'description': action_text,
                'created_at': log_activity['created_at']
            })
        else:
            activities.append({
                'user_name': log_activity['user_name'],
                'action_type': 'updated',
                'old_value': None,
                'new_value': None,
                'description': action_text,
                'created_at': log_activity['created_at']
            })
    
    # Sort activities by timestamp
    activities = sorted(activities, key=lambda x: x['created_at'])
    
    # Get attachments from database
    cur.execute("""
        SELECT 
            ta.id, ta.file_name, ta.file_path, ta.file_size, 
            ta.file_type, ta.uploaded_at,
            CONCAT(m.first_name, ' ', m.last_name) AS uploaded_by_name
        FROM task_attachments ta
        LEFT JOIN members m ON ta.uploaded_by = m.id
        WHERE ta.task_id = %s
        ORDER BY ta.uploaded_at DESC
    """, (task_id,))
    
    attachments = cur.fetchall()
    
    # Convert file sizes to human-readable format
    for attachment in attachments:
        size_bytes = attachment.get('file_size', 0)
        if size_bytes < 1024:
            attachment['file_size_display'] = f"{size_bytes} B"
        elif size_bytes < 1024 * 1024:
            attachment['file_size_display'] = f"{size_bytes / 1024:.1f} KB"
        else:
            attachment['file_size_display'] = f"{size_bytes / (1024 * 1024):.1f} MB"
        
        # Determine file icon based on file type
        file_type = attachment.get('file_type', '').lower()
        if 'image' in file_type:
            attachment['icon'] = 'fa-file-image'
        elif 'pdf' in file_type:
            attachment['icon'] = 'fa-file-pdf'
        elif 'word' in file_type or 'document' in file_type:
            attachment['icon'] = 'fa-file-word'
        elif 'excel' in file_type or 'spreadsheet' in file_type:
            attachment['icon'] = 'fa-file-excel'
        elif 'text' in file_type:
            attachment['icon'] = 'fa-file-alt'
        else:
            attachment['icon'] = 'fa-file'
    
    # Get user role for permissions
    user_role = request.session.get('role', 'USER')
    user_ldap = request.session.get('user_email', '')
    
    cur.close()
    
    return render(
        request,
        'core/task_page_view.html',
        {
            'issue': issue_data,
            'comments': comments,
            'activities': activities,
            'attachments': attachments,
            'user_role': user_role,
            'user_ldap': user_ldap,
            'page': 'task_detail'
        }
    )

# ==============================
#  UPDATE TASK STATUS (API)
# ==============================
@require_POST
def update_task_status(request, task_id):
    '''Update task status via AJAX'''
    import json
    
    try:
        data = json.loads(request.body)
        new_status = data.get('status')
        
        if not new_status:
            return JsonResponse({'success': False, 'error': 'Status is required'})
        
        conn = get_tenant_conn(request)
        cur = conn.cursor()
        
        # Get old status
        cur.execute('SELECT status FROM tasks WHERE id = %s', (task_id,))
        task = cur.fetchone()
        old_status = task['status'] if task else None
        
        # Update task status
        cur.execute('''
            UPDATE tasks 
            SET status = %s, updated_at = NOW()
            WHERE id = %s
        ''', (new_status, task_id))
        
        # Log activity
        user_id = request.session.get('user_id')
        cur.execute('''
            INSERT INTO activity_log 
            (entity_type, entity_id, action, performed_by, timestamp)
            VALUES ('task', %s, %s, %s, NOW())
        ''', (task_id, f'Changed status from {old_status} to {new_status}', user_id))
        
        conn.commit()
        cur.close()
        
        return JsonResponse({'success': True, 'message': 'Status updated successfully'})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# ==============================
#  UPDATE TASK PRIORITY (API)
# ==============================
@require_POST
def update_task_priority(request, task_id):
    '''Update task priority via AJAX'''
    import json
    
    try:
        data = json.loads(request.body)
        new_priority = data.get('priority')
        
        if not new_priority:
            return JsonResponse({'success': False, 'error': 'Priority is required'})
        
        conn = get_tenant_conn(request)
        cur = conn.cursor()
        
        # Get old priority
        cur.execute('SELECT priority FROM tasks WHERE id = %s', (task_id,))
        task = cur.fetchone()
        old_priority = task['priority'] if task else None
        
        # Update task priority
        cur.execute('''
            UPDATE tasks 
            SET priority = %s, updated_at = NOW()
            WHERE id = %s
        ''', (new_priority, task_id))
        
        # Log activity
        user_id = request.session.get('user_id')
        cur.execute('''
            INSERT INTO activity_log 
            (entity_type, entity_id, action, performed_by, timestamp)
            VALUES ('task', %s, %s, %s, NOW())
        ''', (task_id, f'Changed priority from {old_priority} to {new_priority}', user_id))
        
        conn.commit()
        cur.close()
        
        return JsonResponse({'success': True, 'message': 'Priority updated successfully'})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# ==============================
#  ADD TASK COMMENT (API)
# ==============================
@require_POST
def add_task_comment(request, task_id):
    '''Add comment to task via AJAX'''
    import json
    
    try:
        data = json.loads(request.body)
        comment_text = data.get('comment_text', '').strip()
        is_internal = data.get('is_internal', False)
        
        if not comment_text:
            return JsonResponse({'success': False, 'error': 'Comment text is required'})
        
        conn = get_tenant_conn(request)
        cur = conn.cursor()
        
        # Get commenter info
        user_id = request.session.get('user_id')
        user_email = request.session.get('user_email', '')
        
        # Get commenter name
        cur.execute('''
            SELECT CONCAT(first_name, ' ', last_name) as name 
            FROM members 
            WHERE id = %s
        ''', (user_id,))
        
        commenter = cur.fetchone()
        commenter_name = commenter['name'] if commenter else user_email
        
        # Insert comment
        cur.execute('''
            INSERT INTO task_comments 
            (task_id, comment_text, commenter_id, commenter_name, is_internal, created_at)
            VALUES (%s, %s, %s, %s, %s, NOW())
        ''', (task_id, comment_text, user_id, commenter_name, is_internal))
        
        # Log activity
        comment_type = "internal comment" if is_internal else "comment"
        cur.execute('''
            INSERT INTO activity_log 
            (entity_type, entity_id, action, performed_by, timestamp)
            VALUES ('task', %s, %s, %s, NOW())
        ''', (task_id, f'Added {comment_type}', user_id))
        
        conn.commit()
        cur.close()
        
        return JsonResponse({'success': True, 'message': 'Comment posted successfully'})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_POST
def upload_task_attachment(request, task_id):
    '''Upload attachment to task via AJAX'''
    import logging
    logger = logging.getLogger(__name__)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})

    try:
        files = request.FILES.getlist('files')
        logger.info(f"Upload attempt for task {task_id}, files received: {len(files)}")

        if not files:
            logger.warning(f"No files in request for task {task_id}")
            return JsonResponse({'success': False, 'error': 'No files uploaded'})

        conn = get_tenant_conn(request)
        cur = conn.cursor()

        # Verify task exists
        cur.execute('SELECT id FROM tasks WHERE id = %s', (task_id,))
        task_exists = cur.fetchone()
        if not task_exists:
            cur.close()
            logger.error(f"Task {task_id} not found")
            return JsonResponse({'success': False, 'error': 'Task not found'})

        # Get user ID
        user_id = request.session.get('user_id')
        logger.info(f"User ID: {user_id}")

        # Create attachments directory if it doesn't exist
        attachments_dir = os.path.join(settings.MEDIA_ROOT, 'task_attachments')
        os.makedirs(attachments_dir, exist_ok=True)
        logger.info(f"Attachments directory: {attachments_dir}")

        saved_count = 0
        errors = []
        
        for uploaded_file in files:
            try:
                # Generate unique filename with timestamp
                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                safe_filename = f"{task_id}_{timestamp}_{uploaded_file.name}"
                file_path = os.path.join(attachments_dir, safe_filename)
                
                logger.info(f"Saving file: {safe_filename}")

                # Save the file to disk
                with open(file_path, 'wb+') as destination:
                    for chunk in uploaded_file.chunks():
                        destination.write(chunk)

                # Store relative path for database (use forward slashes for consistency)
                relative_path = f'task_attachments/{safe_filename}'

                # Save attachment info to database
                cur.execute("""
                    INSERT INTO task_attachments
                    (task_id, file_name, file_path, file_size, file_type, uploaded_by)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, (
                    task_id,
                    uploaded_file.name,
                    relative_path,
                    uploaded_file.size,
                    uploaded_file.content_type or 'application/octet-stream',
                    user_id
                ))

                # Log attachment upload to activity timeline
                try:
                    cur.execute(
                        "INSERT INTO activity_log (entity_type, entity_id, action, performed_by, timestamp) VALUES (%s,%s,%s,%s,NOW())",
                        ("task", task_id, f"Added attachment {uploaded_file.name}", user_id),
                    )
                except Exception as log_err:
                    logger.warning(f"Failed to log activity: {log_err}")

                saved_count += 1
                logger.info(f"Successfully saved: {safe_filename}")
                
            except Exception as e:
                error_msg = f"Error saving {uploaded_file.name}: {str(e)}"
                logger.error(error_msg)
                errors.append(error_msg)
                continue

        conn.commit()
        cur.close()

        if saved_count > 0:
            logger.info(f"Upload complete: {saved_count} files saved")
            return JsonResponse({
                'success': True,
                'message': f'{saved_count} file(s) uploaded successfully'
            })
        else:
            error_detail = '; '.join(errors) if errors else 'Unknown error'
            logger.error(f"Upload failed: {error_detail}")
            return JsonResponse({
                'success': False,
                'error': f'Failed to upload files: {error_detail}'
            })

    except Exception as e:
        logger.exception(f"Upload exception for task {task_id}")
        return JsonResponse({'success': False, 'error': f'Server error: {str(e)}'})



# ==============================
#  ASSIGN MEMBER TO TASK (API)
# ==============================
@require_POST
def assign_member_to_task(request, task_id):
    '''Assign a member to task via AJAX'''
    import json
    
    try:
        data = json.loads(request.body)
        member_id = data.get('member_id')
        
        if not member_id:
            return JsonResponse({'success': False, 'error': 'Member ID is required'})
        
        conn = get_tenant_conn(request)
        cur = conn.cursor()
        
        # Update task with member assignment
        cur.execute('''
            UPDATE tasks 
            SET assigned_type = 'member',
                assigned_to = %s,
                updated_at = NOW()
            WHERE id = %s
        ''', (member_id, task_id))
        # Log activity: assignment
        try:
            performed_by = request.session.get('user_id')
            cur.execute(
                "INSERT INTO activity_log (entity_type, entity_id, action, performed_by, timestamp) VALUES (%s,%s,%s,%s,NOW())",
                ("task", task_id, f"Assigned to member {member_id}", performed_by),
            )
        except Exception:
            pass
        
        conn.commit()
        cur.close()
        
        return JsonResponse({'success': True, 'message': 'Member assigned successfully'})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# ==============================
#  TASK ANALYTICS - WORK TYPE TRACKING
# ==============================
def task_analytics_view(request):
    """
    Interactive task analytics page showing tasks grouped by work type (Bug, Story, Defect, etc.)
    with filtering capabilities and totals display.
    """
    conn = get_tenant_conn(request)
    cur = conn.cursor()
    
    # Get current user ID
    user_id = request.session.get("user_id")
    if not user_id:
        user_id = request.session.get("member_id")
    
    # Get visible task user IDs based on visibility rules
    visible_user_ids = get_visible_task_user_ids(conn, user_id) if user_id else []
    
    # Get tenant-specific work types
    work_types = get_tenant_work_types(request)
    
    # Initialize data structure for each work type
    work_type_data = {}
    total_counts = {
        'total_tasks': 0,
        'open': 0,
        'in_progress': 0,
        'closed': 0,
        'blocked': 0
    }
    
    # Query tasks grouped by work type and status
    if visible_user_ids:
        placeholders = ','.join(['%s'] * len(visible_user_ids))
        
        # Get all tasks with work type information
        cur.execute(f"""
            SELECT 
                COALESCE(t.work_type, 'Task') AS work_type,
                t.id,
                t.title,
                t.status,
                t.priority,
                t.due_date,
                t.closure_date,
                t.created_at,
                t.assigned_to,
                CONCAT(m.first_name, ' ', m.last_name) AS assigned_name
            FROM tasks t
            LEFT JOIN members m ON m.id = t.assigned_to
            WHERE t.assigned_type='member' AND t.assigned_to IN ({placeholders})
            ORDER BY work_type, FIELD(t.status,'Open','In Progress','Review','Blocked','Closed'), t.created_at DESC
        """, tuple(visible_user_ids))
        
        all_tasks = cur.fetchall()
        
        # Get count statistics by work type
        cur.execute(f"""
            SELECT 
                COALESCE(work_type, 'Task') AS work_type,
                status,
                COUNT(*) AS count
            FROM tasks
            WHERE assigned_type='member' AND assigned_to IN ({placeholders})
            GROUP BY work_type, status
        """, tuple(visible_user_ids))
        
        stats = cur.fetchall()
        
        # Process statistics
        for row in stats:
            wt = row['work_type'] if isinstance(row, dict) else row[0]
            status = row['status'] if isinstance(row, dict) else row[1]
            count = int(row['count'] if isinstance(row, dict) else row[2])
            
            if wt not in work_type_data:
                work_type_data[wt] = {
                    'name': wt,
                    'total': 0,
                    'open': 0,
                    'in_progress': 0,
                    'closed': 0,
                    'blocked': 0,
                    'tasks': []
                }
            
            work_type_data[wt]['total'] += count
            total_counts['total_tasks'] += count
            
            status_lower = status.lower() if status else 'open'
            if status_lower == 'open':
                work_type_data[wt]['open'] += count
                total_counts['open'] += count
            elif status_lower in ('in progress', 'in-progress', 'review'):
                work_type_data[wt]['in_progress'] += count
                total_counts['in_progress'] += count
            elif status_lower == 'closed':
                work_type_data[wt]['closed'] += count
                total_counts['closed'] += count
            elif status_lower == 'blocked':
                work_type_data[wt]['blocked'] += count
                total_counts['blocked'] += count
        
        # Group tasks by work type
        for task in all_tasks:
            wt = task['work_type'] if isinstance(task, dict) else task[0]
            
            if wt not in work_type_data:
                work_type_data[wt] = {
                    'name': wt,
                    'total': 0,
                    'open': 0,
                    'in_progress': 0,
                    'closed': 0,
                    'blocked': 0,
                    'tasks': []
                }
            
            work_type_data[wt]['tasks'].append(task)
    
    cur.close()
    
    return render(
        request,
        "core/task_analytics.html",
        {
            "work_type_data": work_type_data,
            "work_types": work_types,
            "total_counts": total_counts,
            "page": "task_analytics",
            "today": datetime.date.today()
        }
    )
