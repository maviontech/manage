"""
Excel Export functionality for Projects Report
"""
from django.http import HttpResponse
from django.shortcuts import redirect
from io import BytesIO
from datetime import datetime
from .db_helpers import get_tenant_conn

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def export_projects_excel(request):
    """
    Export projects or tasks data to Excel with professional styling
    """
    if not OPENPYXL_AVAILABLE:
        return HttpResponse("openpyxl library is not installed. Please install it using: pip install openpyxl", status=500)
    
    user = request.session.get('user')
    if not user:
        return redirect('login_password')

    member_id = request.session.get('member_id')
    if not member_id:
        return redirect('login_password')

    view_mode = request.GET.get('view_mode', 'projects')
    
    conn = get_tenant_conn(request)
    wb = Workbook()
    ws = wb.active
    
    try:
        # Define professional colors
        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )
        
        if view_mode == 'tasks':
            # Export Tasks
            ws.title = "Tasks Report"
            
            # Headers
            headers = ['ID', 'Task Title', 'Description', 'Priority', 'Status', 'Due Date', 
                      'Assigned To', 'Project', 'Created At', 'Created By']
            ws.append(headers)
            
            # Style headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Fetch tasks data
            cur = conn.cursor()
            cur.execute("""
                SELECT 
                    t.id,
                    t.title,
                    t.description,
                    t.status,
                    t.priority,
                    t.due_date,
                    t.created_at,
                    t.assigned_type,
                    t.assigned_to,
                    p.name AS project_name,
                    m.first_name AS member_first_name,
                    m.last_name AS member_last_name,
                    tm.name AS assigned_team_name,
                    creator.full_name AS created_by_name
                FROM tasks t
                LEFT JOIN projects p ON t.project_id = p.id
                LEFT JOIN members m ON t.assigned_type = 'member' AND t.assigned_to = m.id
                LEFT JOIN teams tm ON t.assigned_type = 'team' AND t.assigned_to = tm.id
                LEFT JOIN users creator ON t.created_by = creator.id
                ORDER BY t.created_at DESC
            """)
            rows = cur.fetchall()
            
            # Add data rows
            for row_num, r in enumerate(rows, 2):
                # Determine assigned name
                assigned_name = 'Unassigned'
                if r['assigned_type'] == 'member' and r['member_first_name']:
                    assigned_name = f"{r['member_first_name']} {r['member_last_name'] or ''}".strip()
                elif r['assigned_type'] == 'team' and r['assigned_team_name']:
                    assigned_name = f"Team: {r['assigned_team_name']}"
                
                row_data = [
                    r['id'],
                    r['title'],
                    r['description'] or '',
                    r['priority'],
                    r['status'],
                    r['due_date'].strftime('%Y-%m-%d') if r['due_date'] else '',
                    assigned_name,
                    r['project_name'] or 'No Project',
                    r['created_at'].strftime('%Y-%m-%d %H:%M') if r['created_at'] else '',
                    r['created_by_name'] or 'Unknown'
                ]
                
                ws.append(row_data)
                
                # Style data rows
                for col_num in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.border = border
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                    
                    # Color code priority
                    if col_num == 4:  # Priority column
                        if r['priority'] == 'Critical':
                            cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
                            cell.font = Font(color='991B1B', bold=True)
                        elif r['priority'] == 'High':
                            cell.fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
                            cell.font = Font(color='92400E', bold=True)
                    
                    # Color code status
                    if col_num == 5:  # Status column
                        if r['status'] == 'Completed':
                            cell.fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
                            cell.font = Font(color='065F46', bold=True)
                        elif r['status'] == 'In Progress':
                            cell.fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
                            cell.font = Font(color='1E40AF', bold=True)
            
            filename = f"Trackline_Tasks_Report_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
            
        else:
            # Export Projects
            ws.title = "Projects Report"
            
            # Headers
            headers = ['Project Name', 'Employee', 'Employee Code', 'Department', 'Designation',
                      'Status', 'Timeline Status', 'Total Tasks', 'Completed Tasks', 'Pending Tasks',
                      'Overdue Tasks', 'Progress %', 'Start Date', 'Due Date', 'End Date', 'Created By']
            ws.append(headers)
            
            # Style headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Fetch projects data
            cur = conn.cursor()
            cur.execute("""
                SELECT 
                    p.id,
                    p.name,
                    p.status,
                    p.start_date,
                    p.tentative_end_date,
                    p.end_date,
                    e.employee_code,
                    e.first_name AS emp_first_name,
                    e.last_name AS emp_last_name,
                    e.department,
                    e.designation,
                    u.full_name AS created_by_name,
                    (SELECT COUNT(*) FROM tasks WHERE project_id = p.id) AS total_tasks,
                    (SELECT COUNT(*) FROM tasks WHERE project_id = p.id AND status = 'Completed') AS completed_tasks,
                    (SELECT COUNT(*) FROM tasks WHERE project_id = p.id AND status IN ('New', 'In Progress', 'Pending')) AS pending_tasks,
                    (SELECT COUNT(*) FROM tasks WHERE project_id = p.id AND due_date < CURDATE() AND status != 'Completed') AS overdue_tasks
                FROM projects p
                LEFT JOIN employees e ON p.employee_id = e.id
                LEFT JOIN users u ON p.created_by = u.id
                ORDER BY p.created_at DESC
            """)
            rows = cur.fetchall()
            
            # Add data rows
            for row_num, r in enumerate(rows, 2):
                # Calculate progress
                total = r['total_tasks'] or 0
                completed = r['completed_tasks'] or 0
                progress = round((completed / total * 100) if total > 0 else 0, 1)
                
                # Calculate timeline status
                timeline_status = 'On Track'
                if r['status'] == 'Completed':
                    timeline_status = 'Completed'
                elif r['tentative_end_date']:
                    from datetime import date
                    today = date.today()
                    end_date = r['tentative_end_date']
                    if isinstance(end_date, str):
                        from datetime import datetime as dt
                        end_date = dt.strptime(end_date, '%Y-%m-%d').date()
                    
                    if end_date < today and r['status'] != 'Completed':
                        timeline_status = 'Overdue'
                    elif (end_date - today).days <= 7 and r['status'] != 'Completed':
                        timeline_status = 'At Risk'
                
                # Employee name
                employee_name = ''
                if r['emp_first_name']:
                    employee_name = f"{r['emp_first_name']} {r['emp_last_name'] or ''}".strip()
                
                row_data = [
                    r['name'],
                    employee_name or 'Not Assigned',
                    r['employee_code'] or 'N/A',
                    r['department'] or 'N/A',
                    r['designation'] or 'N/A',
                    r['status'],
                    timeline_status,
                    total,
                    completed,
                    r['pending_tasks'] or 0,
                    r['overdue_tasks'] or 0,
                    progress,
                    r['start_date'].strftime('%Y-%m-%d') if r['start_date'] else '',
                    r['tentative_end_date'].strftime('%Y-%m-%d') if r['tentative_end_date'] else '',
                    r['end_date'].strftime('%Y-%m-%d') if r['end_date'] else '',
                    r['created_by_name'] or 'Unknown'
                ]
                
                ws.append(row_data)
                
                # Style data rows
                for col_num in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.border = border
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                    
                    # Color code status
                    if col_num == 6:  # Status column
                        if r['status'] == 'Completed':
                            cell.fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
                            cell.font = Font(color='065F46', bold=True)
                        elif r['status'] == 'Active':
                            cell.fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
                            cell.font = Font(color='1E40AF', bold=True)
                    
                    # Color code timeline status
                    if col_num == 7:  # Timeline Status column
                        if timeline_status == 'Overdue':
                            cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
                            cell.font = Font(color='991B1B', bold=True)
                        elif timeline_status == 'At Risk':
                            cell.fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
                            cell.font = Font(color='92400E', bold=True)
                        elif timeline_status == 'Completed':
                            cell.fill = PatternFill(start_color='E0E7FF', end_color='E0E7FF', fill_type='solid')
                            cell.font = Font(color='4338CA', bold=True)
                    
                    # Progress percentage
                    if col_num == 12:
                        cell.number_format = '0.0"%"'
            
            filename = f"Trackline_Projects_Report_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Set row heights
        ws.row_dimensions[1].height = 25
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        print(f"Error exporting to Excel: {e}")
        import traceback
        traceback.print_exc()
        return HttpResponse(f"Error exporting data: {str(e)}", status=500)
    finally:
        if conn:
            conn.close()
